"""
Structural & Semantic Quality Metrics for Spreadsheet Normalization

Usage:
    from structural_metrics import TableQualityMetrics

    metrics = TableQualityMetrics()

    # Compute all deterministic metrics (no LLM/embedding needed)
    report = metrics.compute_structural(df)

    # Compute semantic metrics (requires embedding function)
    from openai import OpenAI
    client = OpenAI()
    def embed_fn(texts):
        resp = client.embeddings.create(model="text-embedding-3-small", input=texts)
        return [e.embedding for e in resp.data]
    sem_report = metrics.compute_semantic(df, embed_fn, dim_cols=["col_a", "col_b"])

    # Full report
    full = metrics.compute_all(df, embed_fn, dim_cols=[...], value_col="count")

    # Before/after comparison
    comparison = metrics.compare(df_before, df_after, dim_cols=[...], value_col="count")
"""

import numpy as np
import pandas as pd
from typing import Any, Callable, Dict, List, Optional, Tuple


# ====================================================================
# STRUCTURAL METRICS (pure deterministic, no LLM)
# ====================================================================

def empty_cell_ratio(df: pd.DataFrame) -> float:
    """
    Fraction of cells that are empty/NaN in the entire table.

    Tidy tables should be close to 0 (only legitimate NULLs).
    Messy tables have high ratios due to merged cells, metadata rows,
    sparse fills, etc.

    Counts both NaN and empty/whitespace-only strings as empty.

    Returns: float in [0, 1]. Lower is better.
    """
    total = df.shape[0] * df.shape[1]
    if total == 0:
        return 0.0

    empty = 0
    for col in df.columns:
        col_data = df[col]
        # Count NaN
        na_count = col_data.isna().sum()
        # Count empty/whitespace strings (for text-like columns)
        try:
            str_vals = col_data.dropna().astype(str).str.strip()
            empty_str_count = (str_vals == "").sum()
        except Exception:
            empty_str_count = 0
        empty += na_count + empty_str_count

    return float(empty / total)


def row_completeness_variance(df: pd.DataFrame) -> float:
    """
    Variance of non-null cell counts across rows.

    In a tidy table every row has the same number of filled columns,
    so variance ≈ 0. In messy tables, metadata rows might have 1
    filled cell while data rows have 8, causing high variance.

    Returns: float >= 0. Lower is better.
    """
    non_null_per_row = df.notna().sum(axis=1)
    return float(non_null_per_row.var())


def column_completeness_variance(df: pd.DataFrame) -> float:
    """
    Variance of non-null cell counts across columns.

    If some columns are 90% empty (e.g., 'sex' only relevant for
    some rows), the variance is high — indicating granularity
    inconsistency.

    Returns: float >= 0. Lower is better.
    """
    non_null_per_col = df.notna().sum(axis=0)
    return float(non_null_per_col.var())


def column_type_purity(df: pd.DataFrame) -> Dict[str, float]:
    """
    For each column, measure how consistently its values are one type.

    A pure column has all numeric or all text values (score = 1.0).
    A mixed column has some numeric and some text (score closer to 0.5).

    Returns: dict of {column_name: purity_score}.
             Each score in [0.5, 1.0]. Higher is better.
    """
    result = {}
    for col in df.columns:
        non_null = df[col].dropna()
        if len(non_null) == 0:
            result[str(col)] = 1.0
            continue
        numeric_count = pd.to_numeric(non_null, errors="coerce").notna().sum()
        ratio = numeric_count / len(non_null)
        # purity = how dominant the majority type is
        result[str(col)] = float(max(ratio, 1.0 - ratio))
    return result


def column_type_purity_mean(df: pd.DataFrame) -> float:
    """Mean type purity across all columns. 1.0 = perfectly pure."""
    purities = column_type_purity(df)
    if not purities:
        return 1.0
    return float(np.mean(list(purities.values())))


def duplicate_key_ratio(df: pd.DataFrame,
                        dim_cols: Optional[List[str]] = None) -> float:
    """
    Fraction of rows that are duplicates on the dimension columns.

    In a tidy table, dimension columns form a unique key (ratio = 0).
    If ratio > 0, there are rows with identical dimension values —
    likely residual aggregation or failed deduplication.

    Args:
        dim_cols: list of dimension column names. If None, uses all columns.

    Returns: float in [0, 1]. 0 = no duplicates (perfect).
    """
    if dim_cols is None:
        dim_cols = list(df.columns)

    # Filter to only existing columns
    dim_cols = [c for c in dim_cols if c in df.columns]
    if not dim_cols:
        return 0.0

    total = len(df)
    if total == 0:
        return 0.0
    unique = len(df.drop_duplicates(subset=dim_cols))
    return float(1.0 - unique / total)


def header_row_count(df: pd.DataFrame, min_numeric_ratio: float = 0.2) -> int:
    """
    Estimate how many rows at the top are header/metadata rows
    (not data).

    Scans from row 0 downward. A data row has at least `min_numeric_ratio`
    of its non-null values being numeric. The first such row is the
    data start; everything above is header.

    Tidy tables: 0 or 1. Messy tables: 3-6+.
    """
    for i in range(len(df)):
        row = df.iloc[i]
        non_null = row.dropna()
        if len(non_null) == 0:
            continue
        numeric = pd.to_numeric(non_null, errors="coerce").notna().sum()
        if numeric / len(non_null) >= min_numeric_ratio:
            return i
    return len(df)  # entire table is non-numeric


# ====================================================================
# SEMANTIC METRICS (some need embeddings, some are numeric-only)
# ====================================================================

def aggregation_residue_score(
        df: pd.DataFrame,
        dim_cols: List[str],
        value_col: str,
        tolerance: float = 0.01,
) -> Dict[str, Any]:
    """
    Detect residual aggregation rows in the output table.

    A residual aggregation row has NULL in some dimension column(s)
    while other rows with the same non-null dimensions have concrete
    values, AND the sum of those detail rows equals this row's value.

    Args:
        dim_cols: dimension columns that should form the key.
        value_col: the numeric value column to check sums on.
        tolerance: absolute tolerance for sum comparison.

    Returns:
        dict with:
          - "score": float in [0, 1]. 0 = no residue (clean).
          - "residue_count": int, number of suspected aggregation rows.
          - "examples": list of example residue rows (up to 5).
    """
    if value_col not in df.columns:
        return {"score": 0.0, "residue_count": 0, "examples": []}

    dim_cols = [c for c in dim_cols if c in df.columns]
    if not dim_cols:
        return {"score": 0.0, "residue_count": 0, "examples": []}

    residue_indices = []

    for i in range(len(df)):
        row = df.iloc[i]
        null_dims = [c for c in dim_cols if pd.isna(row[c])]
        if not null_dims:
            continue

        # This row has NULL in some dimensions — potential aggregate
        non_null_dims = [c for c in dim_cols if pd.notna(row[c])]
        if not non_null_dims:
            continue

        # Find detail rows that match on non-null dims
        mask = pd.Series([True] * len(df), index=df.index)
        for c in non_null_dims:
            mask &= (df[c] == row[c])

        # Among matching rows, find those that have values in the null dims
        for c in null_dims:
            detail_mask = mask & df[c].notna()
            detail_rows = df[detail_mask]
            if len(detail_rows) > 1:
                detail_sum = pd.to_numeric(
                    detail_rows[value_col], errors="coerce"
                ).sum()
                row_val = pd.to_numeric(
                    pd.Series([row[value_col]]), errors="coerce"
                ).iloc[0]
                if pd.notna(row_val) and abs(detail_sum - row_val) < tolerance:
                    residue_indices.append(i)
                    break

    residue_count = len(residue_indices)
    score = residue_count / len(df) if len(df) > 0 else 0.0

    examples = []
    for idx in residue_indices[:5]:
        examples.append(df.iloc[idx].to_dict())

    return {
        "score": float(score),
        "residue_count": residue_count,
        "examples": examples,
    }


def column_semantic_coherence(
        df: pd.DataFrame,
        col: str,
        embed_fn: Callable[[List[str]], List[List[float]]],
        sample_size: int = 50,
) -> Dict[str, float]:
    """
    Measure how semantically consistent values within a column are.

    Uses embeddings + cosine similarity. A column where all values
    are the same kind of thing (all district names, all abuse types)
    will have high mean similarity and low std.

    Args:
        col: column name to evaluate.
        embed_fn: function that takes a list of strings and returns
                  a list of embedding vectors.
        sample_size: max unique values to sample (controls cost).

    Returns:
        dict with:
          - "mean_similarity": float in [-1, 1]. Higher = more coherent.
          - "std_similarity": float >= 0. Lower = more coherent.
          - "n_unique": number of unique values evaluated.
    """
    if col not in df.columns:
        return {"mean_similarity": 0.0, "std_similarity": 0.0, "n_unique": 0}

    unique_vals = df[col].dropna().unique()
    unique_vals = [str(v).strip() for v in unique_vals if str(v).strip()]
    unique_vals = list(set(unique_vals))  # deduplicate after stripping

    if len(unique_vals) < 2:
        return {"mean_similarity": 1.0, "std_similarity": 0.0,
                "n_unique": len(unique_vals)}

    # Sample if too many
    if len(unique_vals) > sample_size:
        rng = np.random.default_rng(42)
        unique_vals = list(rng.choice(unique_vals, sample_size, replace=False))

    # Get embeddings (single batch call)
    embeddings = np.array(embed_fn(unique_vals))

    # Cosine similarity matrix
    # Normalize rows first
    norms = np.linalg.norm(embeddings, axis=1, keepdims=True)
    norms[norms == 0] = 1.0
    normed = embeddings / norms
    sim_matrix = normed @ normed.T

    # Extract upper triangle (exclude diagonal = self-similarity)
    mask = np.triu(np.ones_like(sim_matrix, dtype=bool), k=1)
    pairwise = sim_matrix[mask]

    return {
        "mean_similarity": float(pairwise.mean()),
        "std_similarity": float(pairwise.std()),
        "n_unique": len(unique_vals),
    }


def cross_column_independence(
        df: pd.DataFrame,
        col_a: str,
        col_b: str,
) -> float:
    """
    Normalized Mutual Information between two columns.

    If NMI ≈ 1.0, one column fully determines the other (redundant).
    If NMI ≈ 0.0, columns are independent.

    Useful for detecting:
    - CN/EN translation pairs (NMI ≈ 1.0, acceptable but flagged)
    - Embedded dimensions not yet split (NMI high between compound
      column and a would-be sub-column)

    Returns: float in [0, 1]. Lower = more independent.
    """
    if col_a not in df.columns or col_b not in df.columns:
        return 0.0

    from sklearn.metrics import normalized_mutual_info_score

    # Drop rows where either is null
    valid = df[[col_a, col_b]].dropna()
    if len(valid) < 2:
        return 0.0

    a = valid[col_a].astype(str).values
    b = valid[col_b].astype(str).values

    return float(normalized_mutual_info_score(a, b))


def column_cardinality_profile(
        df: pd.DataFrame,
        dim_cols: Optional[List[str]] = None,
) -> Dict[str, int]:
    """
    Number of unique values per dimension column.

    Useful for spotting granularity mismatches: if one dim column
    has 3 values and another has 300, they represent very different
    levels of detail.

    Returns: dict of {column_name: unique_count}.
    """
    if dim_cols is None:
        dim_cols = list(df.columns)
    dim_cols = [c for c in dim_cols if c in df.columns]
    return {str(c): int(df[c].nunique()) for c in dim_cols}


# ====================================================================
# HIERARCHY METRICS
# ====================================================================

def merged_cell_count(file_path: str) -> Dict[str, Any]:
    """
    Count merged cell ranges in the spreadsheet.

    Merged cells are a strong signal of multi-level headers and
    hierarchical structure. Tidy tables have 0 merged cells.

    Args:
        file_path: path to .xlsx file.

    Returns:
        dict with:
          - "count": total number of merged ranges.
          - "total_cells_affected": sum of cells covered by merges.
          - "ranges": list of merge range strings (e.g., "A1:C1").
    """
    try:
        import openpyxl
        wb = openpyxl.load_workbook(file_path, data_only=True)
        ws = wb.worksheets[0]
        ranges = list(ws.merged_cells.ranges)
        total_cells = 0
        for mr in ranges:
            total_cells += mr.size["rows"] * mr.size["columns"]
        return {
            "count": len(ranges),
            "total_cells_affected": total_cells,
            "ranges": [str(r) for r in ranges],
        }
    except Exception:
        return {"count": 0, "total_cells_affected": 0, "ranges": []}


def indent_hierarchy_depth(file_path: str, label_col: int = 1) -> Dict[str, Any]:
    """
    Detect hierarchy encoded via cell indentation (alignment.indent).

    Many HK government spreadsheets use Excel's indent feature to
    encode parent-child relationships:
      indent=0: top-level category (e.g., "Asian")
      indent=1: sub-category (e.g., "Filipino", "South Asian")
      indent=2: sub-sub-category (e.g., "Indian", "Nepalese")

    Args:
        file_path: path to .xlsx file.
        label_col: 1-indexed column number to check for indentation.

    Returns:
        dict with:
          - "max_depth": maximum indent level found.
          - "indent_distribution": {indent_level: row_count}.
          - "has_hierarchy": True if max_depth > 0.
          - "sample": list of (row, indent, value) tuples (up to 10).
    """
    try:
        import openpyxl
        wb = openpyxl.load_workbook(file_path, data_only=True)
        ws = wb.worksheets[0]

        indent_dist: Dict[int, int] = {}
        samples = []

        for r in range(1, ws.max_row + 1):
            cell = ws.cell(r, label_col)
            if cell.value is None:
                continue
            indent = int(cell.alignment.indent) if cell.alignment else 0
            indent_dist[indent] = indent_dist.get(indent, 0) + 1
            if len(samples) < 10 and indent > 0:
                samples.append((r, indent, str(cell.value)[:50]))

        max_depth = max(indent_dist.keys()) if indent_dist else 0

        return {
            "max_depth": max_depth,
            "indent_distribution": dict(sorted(indent_dist.items())),
            "has_hierarchy": max_depth > 0,
            "sample": samples,
        }
    except Exception as e:
        return {
            "max_depth": 0,
            "indent_distribution": {},
            "has_hierarchy": False,
            "sample": [],
        }


def whitespace_hierarchy_depth(df: pd.DataFrame,
                               label_col: int = 0) -> Dict[str, Any]:
    """
    Detect hierarchy encoded via leading whitespace (NBSP or spaces).

    Some spreadsheets use leading \\xa0 (non-breaking space) or regular
    spaces to indicate hierarchy level, rather than Excel indent.

    Args:
        label_col: positional column index to check.

    Returns:
        dict with:
          - "max_depth": maximum whitespace indent level found.
          - "indent_distribution": {indent_chars: row_count}.
          - "has_hierarchy": True if any rows have leading whitespace.
          - "sample": list of (row_idx, indent_chars, value) tuples.
    """
    if label_col >= df.shape[1]:
        return {"max_depth": 0, "indent_distribution": {},
                "has_hierarchy": False, "sample": []}

    col_data = df.iloc[:, label_col]
    indent_dist: Dict[int, int] = {}
    samples = []

    for i, val in col_data.items():
        if pd.isna(val):
            continue
        s = str(val)
        stripped = s.lstrip("\xa0").lstrip(" ")
        indent = len(s) - len(stripped)
        indent_dist[indent] = indent_dist.get(indent, 0) + 1
        if len(samples) < 10 and indent > 0:
            samples.append((int(i), indent, s[:50]))

    max_depth = max(indent_dist.keys()) if indent_dist else 0

    return {
        "max_depth": max_depth,
        "indent_distribution": dict(sorted(indent_dist.items())),
        "has_hierarchy": max_depth > 0,
        "sample": samples,
    }


def multi_level_header_depth(file_path: str,
                             data_start_row: Optional[int] = None,
                             ) -> Dict[str, Any]:
    """
    Measure the depth of multi-level column headers.

    Scans rows above the data region to find how many rows
    participate in forming the column headers. Uses merged cells
    and non-empty cell patterns to determine header structure.

    Args:
        file_path: path to .xlsx file.
        data_start_row: 1-indexed row where data starts. If None, auto-detect.

    Returns:
        dict with:
          - "header_depth": number of rows used for headers.
          - "has_merged_headers": whether any header rows have merged cells.
          - "header_rows": list of row indices that are part of the header.
    """
    try:
        import openpyxl
        wb = openpyxl.load_workbook(file_path, data_only=True)
        ws = wb.worksheets[0]

        # Auto-detect data start if not provided
        if data_start_row is None:
            for r in range(1, ws.max_row + 1):
                numeric_count = 0
                non_null = 0
                for c in range(1, ws.max_column + 1):
                    v = ws.cell(r, c).value
                    if v is not None:
                        non_null += 1
                        try:
                            float(v)
                            numeric_count += 1
                        except (ValueError, TypeError):
                            pass
                if non_null > 0 and numeric_count / non_null >= 0.2:
                    data_start_row = r
                    break
            if data_start_row is None:
                data_start_row = ws.max_row

        # Check which rows above data_start are header rows
        merged_ranges = list(ws.merged_cells.ranges)
        header_rows = []
        has_merged = False

        for r in range(1, data_start_row):
            row_has_content = False
            for c in range(1, ws.max_column + 1):
                if ws.cell(r, c).value is not None:
                    row_has_content = True
                    break
            if row_has_content:
                header_rows.append(r)
                # Check if any merged cell spans this row
                for mr in merged_ranges:
                    if mr.min_row <= r <= mr.max_row:
                        has_merged = True
                        break

        return {
            "header_depth": len(header_rows),
            "has_merged_headers": has_merged,
            "header_rows": header_rows,
            "data_start_row": data_start_row,
        }
    except Exception:
        return {
            "header_depth": 0,
            "has_merged_headers": False,
            "header_rows": [],
            "data_start_row": 1,
        }


def hierarchy_score(file_path: str,
                    df: pd.DataFrame,
                    label_col: int = 0) -> Dict[str, Any]:
    """
    Composite hierarchy metric combining all hierarchy signals.

    Checks three sources of hierarchy:
    1. Merged cells (column header hierarchy)
    2. Excel alignment indent (row label hierarchy)
    3. Leading whitespace (row label hierarchy)

    A tidy table should have hierarchy_score = 0 (completely flat).

    Returns:
        dict with:
          - "score": float >= 0. 0 = flat table. Higher = more hierarchical.
          - "merged_cells": merged cell analysis.
          - "indent_hierarchy": indent-based hierarchy analysis.
          - "whitespace_hierarchy": whitespace-based hierarchy analysis.
          - "multi_level_header": header depth analysis.
    """
    merged = merged_cell_count(file_path)
    indent = indent_hierarchy_depth(file_path, label_col + 1)  # 1-indexed
    whitespace = whitespace_hierarchy_depth(df, label_col)
    header = multi_level_header_depth(file_path)

    # Composite score: sum of all hierarchy signals
    score = (
            min(merged["count"], 10) / 10 * 0.25           # merged cells
            + min(indent["max_depth"], 5) / 5 * 0.25       # indent depth
            + min(whitespace["max_depth"], 5) / 5 * 0.25   # whitespace depth
            + min(header["header_depth"], 5) / 5 * 0.25    # header depth
    )

    return {
        "score": float(score),
        "merged_cells": merged,
        "indent_hierarchy": indent,
        "whitespace_hierarchy": whitespace,
        "multi_level_header": header,
    }


# ====================================================================
# COMPOSITE REPORT
# ====================================================================

class TableQualityMetrics:
    """Compute and compare table quality metrics."""

    def compute_structural(self, df: pd.DataFrame,
                           dim_cols: Optional[List[str]] = None,
                           ) -> Dict[str, Any]:
        """Compute all deterministic structural metrics."""
        report = {
            "shape": {"rows": df.shape[0], "cols": df.shape[1]},
            "empty_cell_ratio": empty_cell_ratio(df),
            "row_completeness_variance": row_completeness_variance(df),
            "column_completeness_variance": column_completeness_variance(df),
            "column_type_purity_mean": column_type_purity_mean(df),
            "column_type_purity_detail": column_type_purity(df),
            "header_row_count": header_row_count(df),
            "column_cardinality": column_cardinality_profile(df, dim_cols),
        }
        if dim_cols:
            report["duplicate_key_ratio"] = duplicate_key_ratio(df, dim_cols)
        return report

    def compute_semantic(
            self,
            df: pd.DataFrame,
            embed_fn: Callable[[List[str]], List[List[float]]],
            dim_cols: List[str],
            value_col: Optional[str] = None,
            sample_size: int = 50,
    ) -> Dict[str, Any]:
        """Compute semantic metrics (requires embedding function)."""
        report = {}

        # Column coherence for each dimension column
        coherence = {}
        for col in dim_cols:
            if col in df.columns:
                coherence[col] = column_semantic_coherence(
                    df, col, embed_fn, sample_size
                )
        report["column_coherence"] = coherence

        # Mean coherence across all dim columns
        mean_sims = [v["mean_similarity"] for v in coherence.values()
                     if v["n_unique"] >= 2]
        report["mean_column_coherence"] = (
            float(np.mean(mean_sims)) if mean_sims else 1.0
        )

        # Cross-column independence (all pairs of dim cols)
        independence = {}
        for i, a in enumerate(dim_cols):
            for b in dim_cols[i + 1:]:
                if a in df.columns and b in df.columns:
                    key = f"{a} × {b}"
                    independence[key] = cross_column_independence(df, a, b)
        report["cross_column_independence"] = independence

        # Aggregation residue
        if value_col and value_col in df.columns:
            report["aggregation_residue"] = aggregation_residue_score(
                df, dim_cols, value_col
            )

        return report

    def compute_all(
            self,
            df: pd.DataFrame,
            embed_fn: Optional[Callable] = None,
            dim_cols: Optional[List[str]] = None,
            value_col: Optional[str] = None,
    ) -> Dict[str, Any]:
        """Compute all available metrics."""
        report = self.compute_structural(df, dim_cols)

        if embed_fn and dim_cols:
            semantic = self.compute_semantic(
                df, embed_fn, dim_cols, value_col
            )
            report.update(semantic)

        return report

    def compare(
            self,
            df_before: pd.DataFrame,
            df_after: pd.DataFrame,
            dim_cols_after: Optional[List[str]] = None,
            value_col: Optional[str] = None,
            embed_fn: Optional[Callable] = None,
    ) -> Dict[str, Any]:
        """
        Compare metrics before and after normalization.

        Returns a report showing both values and the delta.
        """
        before = self.compute_structural(df_before)
        after = self.compute_structural(df_after, dim_cols_after)

        comparison = {"before": before, "after": after, "delta": {}}

        # Compute deltas for numeric metrics
        numeric_keys = [
            "empty_cell_ratio",
            "row_completeness_variance",
            "column_completeness_variance",
            "column_type_purity_mean",
            "header_row_count",
        ]
        for key in numeric_keys:
            if key in before and key in after:
                b_val = before[key]
                a_val = after[key]
                comparison["delta"][key] = {
                    "before": b_val,
                    "after": a_val,
                    "change": a_val - b_val,
                    "improved": self._is_improved(key, b_val, a_val),
                }

        # Add semantic metrics for after table if embed_fn provided
        if embed_fn and dim_cols_after:
            sem = self.compute_semantic(
                df_after, embed_fn, dim_cols_after, value_col
            )
            comparison["after_semantic"] = sem

        return comparison

    @staticmethod
    def _is_improved(metric_name: str, before: float, after: float) -> bool:
        """Determine if a metric change is an improvement."""
        # Metrics where LOWER is better
        lower_is_better = {
            "empty_cell_ratio",
            "row_completeness_variance",
            "column_completeness_variance",
            "header_row_count",
        }
        # Metrics where HIGHER is better
        higher_is_better = {
            "column_type_purity_mean",
        }

        if metric_name in lower_is_better:
            return after < before
        elif metric_name in higher_is_better:
            return after > before
        return True  # unknown metric, assume no regression

    @staticmethod
    def format_report(report: Dict[str, Any], indent: int = 0) -> str:
        """Pretty-print a metrics report."""
        lines = []
        prefix = "  " * indent

        for key, val in report.items():
            if isinstance(val, dict):
                lines.append(f"{prefix}{key}:")
                lines.append(
                    TableQualityMetrics.format_report(val, indent + 1)
                )
            elif isinstance(val, float):
                lines.append(f"{prefix}{key}: {val:.4f}")
            else:
                lines.append(f"{prefix}{key}: {val}")

        return "\n".join(lines)

# Test
if __name__ == "__main__":
    # Example usage
    from openai import OpenAI

    client = OpenAI()

    def embed_fn(texts):
        resp = client.embeddings.create(model="text-embedding-3-small", input=texts)
        return [e.embedding for e in resp.data]
    df = pd.read_excel("testdata.xlsx")
    metrics = TableQualityMetrics()
    report = metrics.compute_all(df, embed_fn)
    print(metrics.format_report(report))


