"""
efficient_anchors.py

Connected Components-based Structural Anchor Detection for Spreadsheets
Replaces O(R⁴C⁴) heuristic approach with O(RC) linear algorithm

Author: Generated for spreadsheet normalization project
Date: 2025-01-04
"""

import logging
from collections import defaultdict
from typing import Set, List, Tuple, Dict, Optional
from dataclasses import dataclass

try:
    from openpyxl.utils import get_column_letter, column_index_from_string
    import openpyxl
except ImportError:
    raise ImportError("openpyxl is required. Install with: pip install openpyxl")

logger = logging.getLogger(__name__)


# ============================================================================
# Data Structures
# ============================================================================

@dataclass
class SheetRegion:
    """Represents a rectangular region in a spreadsheet"""
    min_row: int
    max_row: int
    min_col: int
    max_col: int

    def area(self) -> int:
        """Calculate the number of cells in this region"""
        return (self.max_row - self.min_row + 1) * (self.max_col - self.min_col + 1)

    def to_dict(self) -> dict:
        """Convert to dictionary for JSON serialization"""
        return {
            "min_row": self.min_row,
            "max_row": self.max_row,
            "min_col": self.min_col,
            "max_col": self.max_col,
        }


# ============================================================================
# Cell Style Utilities
# ============================================================================

def get_fill_color(cell) -> str:
    """
    Extract fill/background color from a cell.

    Args:
        cell: openpyxl cell object

    Returns:
        str: RGB color string or 'none' if no fill color
    """
    try:
        fill = cell.fill
        if fill and fill.patternType == 'solid':
            # Try foreground color first
            if fill.fgColor and fill.fgColor.rgb:
                return str(fill.fgColor.rgb)
            # Fallback to start color
            if fill.start_color and fill.start_color.rgb:
                return str(fill.start_color.rgb)
            # Try indexed color
            if fill.fgColor and fill.fgColor.index:
                return f"index_{fill.fgColor.index}"
    except Exception as e:
        logger.debug(f"Error getting fill color: {e}")
    return 'none'


def has_border(cell) -> bool:
    """
    Check if a cell has any border styling.

    Args:
        cell: openpyxl cell object

    Returns:
        bool: True if cell has at least one border side styled
    """
    try:
        border = cell.border
        return any([
            border.left and border.left.style,
            border.right and border.right.style,
            border.top and border.top.style,
            border.bottom and border.bottom.style
        ])
    except Exception:
        return False


def get_font_signature(cell) -> Tuple:
    """
    Extract font-related properties for similarity comparison.

    Args:
        cell: openpyxl cell object

    Returns:
        tuple: (bold, italic, size, name)
    """
    try:
        font = cell.font
        if font:
            return (
                bool(font.bold),
                bool(font.italic),
                font.sz if font.sz else 11,
                font.name if font.name else 'Calibri'
            )
    except Exception:
        pass
    return (False, False, 11, 'Calibri')


def get_cell_signature(cell) -> Tuple:
    """
    Create a hashable signature for a cell based on its style properties.
    Used for quick similarity comparison in O(1) time.

    Args:
        cell: openpyxl cell object

    Returns:
        tuple: Signature containing (bold, fill_color, has_border, data_type, number_format)
    """
    try:
        # Font properties
        bold = cell.font.bold if cell.font else False

        # Fill color
        fill_color = get_fill_color(cell)

        # Border presence
        border = has_border(cell)

        # Data type
        data_type = type(cell.value).__name__ if cell.value is not None else 'NoneType'

        # Number format
        number_format = cell.number_format if cell.number_format else 'General'

        return (bold, fill_color, border, data_type, number_format)
    except Exception as e:
        logger.debug(f"Error creating cell signature: {e}")
        return ('error', 'none', False, 'none', 'General')


def cells_similar(cell1, cell2, threshold: float = 0.75) -> bool:
    """
    Determine if two cells should belong to the same connected region.
    Uses weighted scoring system for different style attributes.

    The threshold balances precision (avoiding false connections) with
    recall (detecting actual table regions).

    Args:
        cell1: First openpyxl cell
        cell2: Second openpyxl cell
        threshold: Similarity threshold (0.0-1.0), default 0.75

    Returns:
        bool: True if cells are similar enough to be connected
    """
    score = 0.0
    total = 0.0

    # Weight 1: Font bold (2.0 points) - strong table boundary indicator
    try:
        bold1 = cell1.font.bold if cell1.font else False
        bold2 = cell2.font.bold if cell2.font else False
        if bold1 == bold2:
            score += 2.0
    except Exception:
        pass
    total += 2.0

    # Weight 2: Fill color (3.0 points) - very distinctive visual feature
    if get_fill_color(cell1) == get_fill_color(cell2):
        score += 3.0
    total += 3.0

    # Weight 3: Border presence (1.5 points)
    if has_border(cell1) == has_border(cell2):
        score += 1.5
    total += 1.5

    # Weight 4: Data type (2.0 points) - numeric vs text is important
    type1 = type(cell1.value) if cell1.value is not None else type(None)
    type2 = type(cell2.value) if cell2.value is not None else type(None)
    if type1 == type2:
        score += 2.0
    total += 2.0

    # Weight 5: Number format (1.5 points)
    try:
        fmt1 = cell1.number_format if cell1.number_format else 'General'
        fmt2 = cell2.number_format if cell2.number_format else 'General'
        if fmt1 == fmt2:
            score += 1.5
    except Exception:
        pass
    total += 1.5

    return (score / total) >= threshold


# ============================================================================
# Connected Components Algorithm - Core
# ============================================================================

def build_connectivity_graph(sheet, region: Optional[SheetRegion] = None) -> Dict[str, Set[str]]:
    """
    Build an undirected graph where nodes are cells and edges connect
    similar adjacent cells. Complexity: O(RC).

    This is Phase 1 of the algorithm. Each cell is visited exactly once,
    and for each cell we check at most 4 neighbors (right, bottom).

    Args:
        sheet: openpyxl worksheet
        region: Optional SheetRegion to limit scanning area

    Returns:
        dict: Adjacency list {cell_ref: set(neighbor_refs)}
    """
    # Determine scan region
    if region:
        min_r, max_r = region.min_row, region.max_row
        min_c, max_c = region.min_col, region.max_col
    else:
        min_r, max_r = 1, sheet.max_row
        min_c, max_c = 1, sheet.max_column

    graph = defaultdict(set)

    logger.info(f"Building connectivity graph for region R[{min_r}:{max_r}] x C[{min_c}:{max_c}]")

    # Single pass through all cells - O(R*C)
    cell_count = 0
    edge_count = 0

    for r in range(min_r, max_r + 1):
        for c in range(min_c, max_c + 1):
            cell = sheet.cell(row=r, column=c)

            # Skip empty cells to reduce graph size
            if cell.value is None or (isinstance(cell.value, str) and cell.value.strip() == ''):
                continue

            cell_ref = f"{get_column_letter(c)}{r}"
            cell_count += 1

            # Check right neighbor (same row, next column)
            if c < max_c:
                right_cell = sheet.cell(row=r, column=c + 1)
                right_val = right_cell.value

                # Only connect non-empty similar cells
                if right_val is not None and \
                        not (isinstance(right_val, str) and right_val.strip() == '') and \
                        cells_similar(cell, right_cell):
                    right_ref = f"{get_column_letter(c + 1)}{r}"
                    graph[cell_ref].add(right_ref)
                    graph[right_ref].add(cell_ref)  # Undirected edge
                    edge_count += 1

            # Check bottom neighbor (next row, same column)
            if r < max_r:
                bottom_cell = sheet.cell(row=r + 1, column=c)
                bottom_val = bottom_cell.value

                if bottom_val is not None and \
                        not (isinstance(bottom_val, str) and bottom_val.strip() == '') and \
                        cells_similar(cell, bottom_cell):
                    bottom_ref = f"{get_column_letter(c)}{r + 1}"
                    graph[cell_ref].add(bottom_ref)
                    graph[bottom_ref].add(cell_ref)  # Undirected edge
                    edge_count += 1

    logger.info(f"Graph built: {cell_count} nodes, {edge_count} edges")
    return dict(graph)


def find_connected_components(graph: Dict[str, Set[str]],
                              min_component_size: int = 4) -> List[List[str]]:
    """
    Find all connected components using Depth-First Search (DFS).
    Complexity: O(V+E) where V=nodes, E=edges. Since E≤4V, this is O(V)=O(RC).

    This is Phase 2 of the algorithm. Each node and edge is visited exactly once.

    Args:
        graph: Adjacency list representation
        min_component_size: Minimum number of cells to form a valid component
                          (filters out noise and singleton cells)

    Returns:
        list: List of components, where each component is a list of cell references
    """
    visited = set()
    components = []

    logger.info(f"Finding connected components (min size: {min_component_size})")

    # Process each unvisited node - total O(V+E)
    for start_ref in graph:
        if start_ref in visited:
            continue

        # DFS to explore entire component
        component = []
        stack = [start_ref]

        while stack:
            ref = stack.pop()

            # Skip if already visited
            if ref in visited:
                continue

            # Mark as visited and add to component
            visited.add(ref)
            component.append(ref)

            # Add all unvisited neighbors to stack
            for neighbor in graph.get(ref, []):
                if neighbor not in visited:
                    stack.append(neighbor)

        # Filter out components that are too small (likely noise)
        if len(component) >= min_component_size:
            components.append(component)
            logger.debug(f"Found component with {len(component)} cells")

    logger.info(f"Found {len(components)} valid components")

    # Sort by size (largest first) for better anchor selection
    components.sort(key=len, reverse=True)

    return components


def extract_component_boundaries(components: List[List[str]]) -> Tuple[Set[int], Set[int]]:
    """
    Extract row and column boundaries from connected components.
    Complexity: O(total_cells) since we visit each cell in each component once.

    This is Phase 3 of the algorithm. For each component, we find its bounding box
    and extract the four boundaries (top, bottom, left, right) as structural anchors.

    Args:
        components: List of cell reference lists

    Returns:
        tuple: (row_boundaries, col_boundaries) as sets of integers
    """
    row_boundaries = set()
    col_boundaries = set()

    logger.info(f"Extracting boundaries from {len(components)} components")

    for i, component in enumerate(components):
        if not component:
            continue

        # Find bounding box for this component - O(component_size)
        min_r = min_c = float('inf')
        max_r = max_c = 0

        for cell_ref in component:
            col_letter, row = split_cell_ref(cell_ref)
            col = column_index_from_string(col_letter)

            min_r = min(min_r, row)
            max_r = max(max_r, row)
            min_c = min(min_c, col)
            max_c = max(max_c, col)

        # Add all four boundaries as structural anchors
        row_boundaries.add(min_r)  # Top boundary
        row_boundaries.add(max_r)  # Bottom boundary
        col_boundaries.add(min_c)  # Left boundary
        col_boundaries.add(max_c)  # Right boundary

        logger.debug(f"Component {i+1}: R[{min_r}:{max_r}] x C[{min_c}:{max_c}] "
                     f"({len(component)} cells)")

    logger.info(f"Extracted {len(row_boundaries)} row boundaries, "
                f"{len(col_boundaries)} col boundaries")

    return row_boundaries, col_boundaries


def extract_k_neighborhood(indices: Set[int], k: int, max_index: int) -> List[int]:
    """
    Expand index set with k-neighborhood to include headers, titles, and notes
    that are near table boundaries.

    Complexity: O(|indices| * k), which is typically O(num_anchors * k).
    Since num_anchors << R, this is much faster than O(R²).

    This is Phase 4 of the algorithm, implementing the paper's recommendation
    to preserve k rows/columns around each structural anchor.

    Args:
        indices: Set of anchor indices (row or column numbers)
        k: Neighborhood radius (typically 2-4)
        max_index: Maximum valid index (sheet.max_row or sheet.max_column)

    Returns:
        list: Sorted list of expanded indices within valid range [1, max_index]
    """
    expanded = set()

    for idx in indices:
        # Add k cells before and after each anchor
        for offset in range(-k, k + 1):
            expanded_idx = idx + offset
            # Ensure within valid sheet boundaries
            if 1 <= expanded_idx <= max_index:
                expanded.add(expanded_idx)

    return sorted(expanded)


# ============================================================================
# Helper Utilities
# ============================================================================

def split_cell_ref(cell_ref: str) -> Tuple[str, int]:
    """
    Split cell reference into column letter and row number.
    Examples: 'A1' -> ('A', 1), 'AA100' -> ('AA', 100)

    Args:
        cell_ref: Cell reference string like 'A1' or 'AA100'

    Returns:
        tuple: (column_letter, row_number)
    """
    col_str = ''.join(filter(str.isalpha, cell_ref))
    row_str = ''.join(filter(str.isdigit, cell_ref))
    return col_str, int(row_str) if row_str else 0


# ============================================================================
# Main API Functions
# ============================================================================

def efficient_find_structural_anchors(sheet,
                                      k: int = 2,
                                      region: Optional[SheetRegion] = None,
                                      min_component_size: int = 4) -> Tuple[List[int], List[int]]:
    """
    Find structural anchors using connected components algorithm.

    TOTAL COMPLEXITY: O(RC) where R=rows, C=columns

    This is a dramatic improvement over the O(R⁴C⁴) heuristic approach used
    in the paper's Appendix C. The algorithm works by:

    1. Building a connectivity graph of similar cells - O(RC)
    2. Finding connected components via DFS - O(RC)
    3. Extracting bounding boxes as anchors - O(RC)
    4. Expanding with k-neighborhood - O(anchors * k)

    The key insight is that tables naturally form connected regions of
    similarly-styled cells, so we can detect them in linear time instead
    of exhaustively trying all possible rectangles.

    Args:
        sheet: openpyxl worksheet object
        k: Neighborhood radius for anchor expansion (typically 2-4)
        region: Optional SheetRegion to limit processing area
        min_component_size: Minimum cells to form valid component (default 4)

    Returns:
        tuple: (row_anchors, col_anchors) as sorted lists of integers

    Example:
        >>> sheet = workbook['Sheet1']
        >>> row_anchors, col_anchors = efficient_find_structural_anchors(sheet, k=2)
        >>> print(f"Found {len(row_anchors)} row anchors, {len(col_anchors)} col anchors")
    """
    logger.info(f"Starting efficient anchor detection (k={k}, min_size={min_component_size})")

    # Phase 1: Build connectivity graph - O(RC)
    graph = build_connectivity_graph(sheet, region)

    # Handle edge case: empty or very sparse sheet
    if not graph or len(graph) < min_component_size:
        logger.warning("Insufficient cells for component analysis, using full sheet boundaries")
        max_r = region.max_row if region else sheet.max_row
        max_c = region.max_col if region else sheet.max_column
        min_r = region.min_row if region else 1
        min_c = region.min_col if region else 1
        return ([min_r, max_r], [min_c, max_c])

    # Phase 2: Find connected components - O(V+E) = O(RC)
    components = find_connected_components(graph, min_component_size)

    # Handle edge case: no valid components found
    if not components:
        logger.warning("No valid components found, using full sheet boundaries")
        max_r = region.max_row if region else sheet.max_row
        max_c = region.max_col if region else sheet.max_column
        min_r = region.min_row if region else 1
        min_c = region.min_col if region else 1
        return ([min_r, max_r], [min_c, max_c])

    # Phase 3: Extract boundaries from components - O(total_cells)
    row_boundaries, col_boundaries = extract_component_boundaries(components)

    # Phase 4: Expand boundaries with k-neighborhood - O(anchors * k)
    max_r = region.max_row if region else sheet.max_row
    max_c = region.max_col if region else sheet.max_column

    row_anchors = extract_k_neighborhood(row_boundaries, k, max_r)
    col_anchors = extract_k_neighborhood(col_boundaries, k, max_c)

    # Log statistics
    coverage_rows = len(row_anchors) / max_r * 100 if max_r > 0 else 0
    coverage_cols = len(col_anchors) / max_c * 100 if max_c > 0 else 0

    logger.info(f"Anchor detection complete:")
    logger.info(f"  - Row anchors: {len(row_anchors)} ({coverage_rows:.1f}% coverage)")
    logger.info(f"  - Col anchors: {len(col_anchors)} ({coverage_cols:.1f}% coverage)")
    logger.info(f"  - Components found: {len(components)}")
    logger.info(f"  - Largest component: {len(components[0])} cells")

    return row_anchors, col_anchors


def hybrid_find_structural_anchors(sheet,
                                   k: int = 2,
                                   region: Optional[SheetRegion] = None,
                                   size_threshold: int = 10000,
                                   fast_method_func=None) -> Tuple[List[int], List[int]]:
    """
    Hybrid strategy: choose algorithm based on table size for optimal performance.

    Strategy:
    - Small tables (<threshold): Use fast CV method (if provided)
    - Large tables (≥threshold): Use connected components O(RC) algorithm

    This gives you the best of both worlds:
    - Fast execution on small tables
    - Scalability on large tables

    Args:
        sheet: openpyxl worksheet
        k: Neighborhood radius
        region: Optional region to process
        size_threshold: Cell count threshold for algorithm selection (default 10,000)
        fast_method_func: Optional fast method function for small tables
                         If None, always uses connected components method

    Returns:
        tuple: (row_anchors, col_anchors) as sorted integer lists

    Example:
        >>> from old_encoder import fast_find_structural_anchors as fast_cv
        >>> row_anchors, col_anchors = hybrid_find_structural_anchors(
        ...     sheet, k=2, size_threshold=10000, fast_method_func=fast_cv
        ... )
    """
    # Calculate table size
    if region:
        num_rows = region.max_row - region.min_row + 1
        num_cols = region.max_col - region.min_col + 1
    else:
        num_rows = sheet.max_row
        num_cols = sheet.max_column

    total_cells = num_rows * num_cols

    logger.info(f"Table size: {num_rows} x {num_cols} = {total_cells} cells")

    # Choose algorithm based on size
    if total_cells < size_threshold and fast_method_func is not None:
        logger.info(f"Using fast method (table < {size_threshold} cells)")
        return fast_method_func(sheet, k, region=region)
    else:
        logger.info(f"Using connected components method (table ≥ {size_threshold} cells)")
        return efficient_find_structural_anchors(sheet, k, region)


# ============================================================================
# Debugging and Visualization Utilities
# ============================================================================

def analyze_components(sheet, k: int = 2, region: Optional[SheetRegion] = None) -> Dict:
    """
    Analyze connected components for debugging and understanding table structure.
    Returns detailed statistics about detected components.

    Args:
        sheet: openpyxl worksheet
        k: Neighborhood radius
        region: Optional region to analyze

    Returns:
        dict: Statistics including component sizes, boundaries, coverage, etc.
    """
    graph = build_connectivity_graph(sheet, region)
    components = find_connected_components(graph, min_component_size=1)  # Include all

    # Calculate statistics
    component_sizes = [len(c) for c in components]
    row_boundaries, col_boundaries = extract_component_boundaries(components)

    max_r = region.max_row if region else sheet.max_row
    max_c = region.max_col if region else sheet.max_column

    stats = {
        'num_components': len(components),
        'component_sizes': component_sizes,
        'largest_component': max(component_sizes) if component_sizes else 0,
        'smallest_component': min(component_sizes) if component_sizes else 0,
        'avg_component_size': sum(component_sizes) / len(component_sizes) if component_sizes else 0,
        'total_cells_in_components': sum(component_sizes),
        'num_row_boundaries': len(row_boundaries),
        'num_col_boundaries': len(col_boundaries),
        'row_coverage_percent': len(row_boundaries) / max_r * 100 if max_r > 0 else 0,
        'col_coverage_percent': len(col_boundaries) / max_c * 100 if max_c > 0 else 0,
    }

    # Add top 5 largest components details
    stats['top_5_components'] = []
    for i, comp in enumerate(components[:5]):
        min_r = min_c = float('inf')
        max_r = max_c = 0
        for cell_ref in comp:
            col_letter, row = split_cell_ref(cell_ref)
            col = column_index_from_string(col_letter)
            min_r = min(min_r, row)
            max_r = max(max_r, row)
            min_c = min(min_c, col)
            max_c = max(max_c, col)

        stats['top_5_components'].append({
            'rank': i + 1,
            'size': len(comp),
            'bounding_box': f"R{min_r}:R{max_r} x C{min_c}:C{max_c}",
            'dimensions': f"{max_r - min_r + 1} x {max_c - min_c + 1}"
        })

    return stats


def print_component_analysis(stats: Dict):
    """
    Pretty-print component analysis results.

    Args:
        stats: Dictionary returned by analyze_components()
    """
    print("\n" + "="*80)
    print("CONNECTED COMPONENTS ANALYSIS")
    print("="*80)
    print(f"Total components found: {stats['num_components']}")
    print(f"Total cells in components: {stats['total_cells_in_components']}")
    print(f"\nComponent sizes:")
    print(f"  - Largest: {stats['largest_component']} cells")
    print(f"  - Smallest: {stats['smallest_component']} cells")
    print(f"  - Average: {stats['avg_component_size']:.1f} cells")
    print(f"\nBoundaries detected:")
    print(f"  - Row boundaries: {stats['num_row_boundaries']} ({stats['row_coverage_percent']:.1f}% coverage)")
    print(f"  - Col boundaries: {stats['num_col_boundaries']} ({stats['col_coverage_percent']:.1f}% coverage)")

    if stats['top_5_components']:
        print(f"\nTop 5 largest components:")
        for comp in stats['top_5_components']:
            print(f"  {comp['rank']}. {comp['size']} cells - {comp['bounding_box']} ({comp['dimensions']})")

    print("="*80 + "\n")


# ============================================================================
# Example Usage and Testing
# ============================================================================

if __name__ == "__main__":
    """
    Example usage and testing code.
    Run this file directly to test the algorithm on your spreadsheet.
    """
    import sys

    # Setup logging
    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
    )

    # Example 1: Basic usage
    print("\n" + "="*80)
    print("EXAMPLE 1: Basic Usage")
    print("="*80)

    if len(sys.argv) > 1:
        excel_file = sys.argv[1]
        print(f"Loading: {excel_file}")

        workbook = openpyxl.load_workbook(excel_file, data_only=True)
        sheet = workbook.active

        print(f"Sheet: {sheet.title}")
        print(f"Dimensions: {sheet.max_row} rows x {sheet.max_column} columns")

        # Run efficient anchor detection
        import time
        start = time.time()
        row_anchors, col_anchors = efficient_find_structural_anchors(sheet, k=2)
        elapsed = time.time() - start

        print(f"\nResults:")
        print(f"  - Execution time: {elapsed:.3f} seconds")
        print(f"  - Row anchors: {len(row_anchors)} (coverage: {len(row_anchors)/sheet.max_row*100:.1f}%)")
        print(f"  - Col anchors: {len(col_anchors)} (coverage: {len(col_anchors)/sheet.max_column*100:.1f}%)")
        print(f"  - First 10 row anchors: {row_anchors[:10]}")
        print(f"  - First 10 col anchors: {col_anchors[:10]}")

        # Example 2: Detailed analysis
        print("\n" + "="*80)
        print("EXAMPLE 2: Detailed Component Analysis")
        print("="*80)

        stats = analyze_components(sheet, k=2)
        print_component_analysis(stats)

    else:
        print("Usage: python efficient_anchors.py <excel_file>")
        print("\nExample:")
        print("  python efficient_anchors.py population.xlsx")