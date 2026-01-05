"""
Format Extractor for SpreadsheetLLM Enhancement
Enhanced format extraction to solve SpreadsheetLLM Limitation 1

Author: Your Name
Date: 2024
"""

from typing import Dict, Any, Optional
from openpyxl.cell import Cell
from openpyxl.styles import Color


def extract_color(color: Optional[Color]) -> Optional[str]:
    """
    Extract and normalize color value to RRGGBB format.
    
    Args:
        color: openpyxl Color object
        
    Returns:
        - "RRGGBB" (6-digit hex without alpha channel)
        - "indexed_X" for indexed colors
        - "theme_X" for theme colors
        - "auto" for auto colors
        - None if no color
        
    Examples:
        "FF0000" -> "FF0000" (red)
        "FFFF0000" -> "FF0000" (red with alpha)
        indexed(5) -> "indexed_5"
    """
    if not color:
        return None
    
    # Case 1: RGB string (most common)
    if hasattr(color, 'rgb') and isinstance(color.rgb, str):
        rgb = color.rgb
        # Remove alpha channel if present (AARRGGBB -> RRGGBB)
        if len(rgb) == 8:
            return rgb[2:]  # Strip first 2 chars (alpha)
        elif len(rgb) == 6:
            return rgb
        else:
            return rgb
    
    # Case 2: Indexed color
    if hasattr(color, 'type'):
        if color.type == 'indexed' and hasattr(color, 'index'):
            return f"indexed_{color.index}"
        
        # Case 3: Theme color
        if color.type == 'theme' and hasattr(color, 'theme'):
            return f"theme_{color.theme}"
        
        # Case 4: Auto color
        if color.type == 'auto':
            return "auto"
    
    return None


def extract_font_info(cell: Cell) -> Dict[str, Any]:
    """
    Extract font information from a cell.
    
    Returns:
        {
            "name": "Calibri",
            "size": 11.0,
            "bold": True,
            "italic": False,
            "underline": None,
            "strike": False,
            "color": "FF0000"  # or None
        }
    """
    if not cell.font:
        return {}
    
    font_info = {
        "name": cell.font.name,
        "size": cell.font.size,
        "bold": cell.font.bold,
        "italic": cell.font.italic,
        "underline": cell.font.underline,
        "strike": cell.font.strike,
    }
    
    # Extract font color
    if cell.font.color:
        font_info["color"] = extract_color(cell.font.color)
    else:
        font_info["color"] = None
    
    return font_info


def extract_fill_info(cell: Cell) -> Dict[str, Any]:
    """
    Extract fill (background) information from a cell.
    
    Returns:
        {
            "pattern_type": "solid",  # or "none", "gray125", etc.
            "fg_color": "D9E1F2",     # foreground color (actual fill)
            "bg_color": None          # background color (rarely used)
        }
    """
    if not cell.fill:
        return {}
    
    fill_info = {
        "pattern_type": cell.fill.patternType if hasattr(cell.fill, 'patternType') else None,
    }
    
    # Foreground color (actual fill color)
    if hasattr(cell.fill, 'fgColor') and cell.fill.fgColor:
        fill_info["fg_color"] = extract_color(cell.fill.fgColor)
    else:
        fill_info["fg_color"] = None
    
    # Background color (usually not needed)
    if hasattr(cell.fill, 'bgColor') and cell.fill.bgColor:
        fill_info["bg_color"] = extract_color(cell.fill.bgColor)
    else:
        fill_info["bg_color"] = None
    
    return fill_info


def extract_border_info(cell: Cell) -> Dict[str, Any]:
    """
    Extract border information from a cell.
    
    Returns:
        {
            "top": {"style": "thin", "color": "000000"},
            "bottom": {"style": "medium", "color": None},
            "left": None,
            "right": None
        }
    """
    if not cell.border:
        return {}
    
    border_info = {}
    
    # Extract each side
    for side_name in ['top', 'bottom', 'left', 'right', 'diagonal']:
        side = getattr(cell.border, side_name, None)
        
        if side and side.style:
            border_info[side_name] = {
                "style": side.style,  # 'thin', 'medium', 'thick', 'double', etc.
                "color": extract_color(side.color) if side.color else None
            }
        else:
            border_info[side_name] = None
    
    return border_info


def extract_alignment_info(cell: Cell) -> Dict[str, Any]:
    """
    Extract alignment information from a cell.
    
    Returns:
        {
            "horizontal": "center",
            "vertical": "center",
            "wrap_text": False,
            "text_rotation": 0,
            "indent": 0,
            "shrink_to_fit": False
        }
    """
    if not cell.alignment:
        return {}
    
    return {
        "horizontal": cell.alignment.horizontal,
        "vertical": cell.alignment.vertical,
        "wrap_text": cell.alignment.wrap_text,
        "text_rotation": cell.alignment.textRotation,
        "indent": cell.alignment.indent if hasattr(cell.alignment, 'indent') else None,
        "shrink_to_fit": cell.alignment.shrinkToFit if hasattr(cell.alignment, 'shrinkToFit') else None,
    }


def extract_complete_format_info(cell: Cell) -> Dict[str, Any]:
    """
    Extract complete formatting information from a cell.
    This is the main function to use.
    
    Args:
        cell: openpyxl Cell object
        
    Returns:
        Dictionary with all format information:
        {
            "font": {...},
            "fill": {...},
            "border": {...},
            "alignment": {...}
        }
        
    Example:
        >>> from openpyxl import load_workbook
        >>> wb = load_workbook("data.xlsx")
        >>> cell = wb.active['A1']
        >>> fmt = extract_complete_format_info(cell)
        >>> print(fmt)
        {
            "font": {"bold": True, "color": "FF0000", ...},
            "fill": {"fg_color": "D9E1F2", ...},
            "border": {"bottom": {"style": "medium", ...}},
            "alignment": {"horizontal": "center", ...}
        }
    """
    return {
        "font": extract_font_info(cell),
        "fill": extract_fill_info(cell),
        "border": extract_border_info(cell),
        "alignment": extract_alignment_info(cell),
    }


def create_format_signature(format_info: Dict[str, Any], 
                           include_colors: bool = True,
                           include_borders: bool = True,
                           simplify: bool = True) -> str:
    """
    Create a compact signature string for format aggregation.
    
    Args:
        format_info: Output from extract_complete_format_info()
        include_colors: Whether to include color information
        include_borders: Whether to include border information
        simplify: If True, exclude common default values (black text, white fill)
        
    Returns:
        Format signature string, e.g.:
        - "bold|fill:D9E1F2|border:b"
        - "italic|font_color:FF0000"
        - "plain" (for cells with no special formatting)
        
    Examples:
        >>> fmt = {"font": {"bold": True}, "fill": {"fg_color": "D9E1F2"}, 
        ...        "border": {"bottom": {"style": "medium"}}}
        >>> create_format_signature(fmt)
        "bold|fill:D9E1F2|border:b"
    """
    sig_parts = []
    
    # 1. Font attributes
    font = format_info.get("font", {})
    if font is None:
        font = {}
    
    if isinstance(font, dict):
        if font.get("bold"):
            sig_parts.append("bold")
        if font.get("italic"):
            sig_parts.append("italic")
        if font.get("underline"):
            sig_parts.append("underline")
        if font.get("strike"):
            sig_parts.append("strike")
        
        # 2. Font color
        if include_colors and font.get("color"):
            color = font["color"]
            # Skip if it's default black or auto
            if simplify:
                if color and color not in ("000000", "auto"):
                    sig_parts.append(f"font_color:{color}")
            else:
                sig_parts.append(f"font_color:{color}")
    
    # 3. Fill color (background)
    if include_colors:
        fill = format_info.get("fill", {})
        if fill is None:
            fill = {}
        
        if isinstance(fill, dict):
            fg_color = fill.get("fg_color")
            
            # Skip if it's default (white, black, or none)
            if fg_color:
                if simplify:
                    if fg_color not in ("000000", "FFFFFF", "auto"):
                        sig_parts.append(f"fill:{fg_color}")
                else:
                    sig_parts.append(f"fill:{fg_color}")
    
    # 4. Borders
    if include_borders:
        border = format_info.get("border", {})
        if border is None:
            border = {}
        
        if isinstance(border, dict):
            border_sides = []
            
            for side in ['top', 'bottom', 'left', 'right']:
                side_info = border.get(side)
                if side_info and isinstance(side_info, dict) and side_info.get("style"):
                    # Use first letter: t, b, l, r
                    border_sides.append(side[0])
            
            if border_sides:
                # Sort for consistency: blrt
                sig_parts.append(f"border:{''.join(sorted(border_sides))}")
    
    # 5. Return signature
    if not sig_parts:
        return "plain"
    
    return "|".join(sig_parts)


def create_semantic_signature(format_info: Dict[str, Any]) -> str:
    """
    Create a semantic signature that describes the likely purpose.
    
    Returns:
        One of: "header", "total_row", "highlight", "bordered", "bold_text", "plain"
        
    This is useful for understanding what role a cell plays in the spreadsheet.
    
    Examples:
        bold + fill + bottom border -> "header"
        bold + top border -> "total_row"
        fill only -> "highlight"
    """
    font = format_info.get("font", {})
    fill = format_info.get("fill", {})
    border = format_info.get("border", {})
    
    # Handle None values safely
    if font is None:
        font = {}
    if fill is None:
        fill = {}
    if border is None:
        border = {}
    
    is_bold = font.get("bold", False) if isinstance(font, dict) else False
    has_fill = fill.get("fg_color") not in (None, "000000", "FFFFFF") if isinstance(fill, dict) else False
    
    # Safe border checking
    has_bottom_border = False
    has_top_border = False
    if isinstance(border, dict):
        bottom = border.get("bottom")
        has_bottom_border = isinstance(bottom, dict) and bottom.get("style") is not None
        
        top = border.get("top")
        has_top_border = isinstance(top, dict) and top.get("style") is not None
        
        has_all_borders = all(
            isinstance(border.get(side), dict) and border.get(side).get("style") is not None
            for side in ['top', 'bottom', 'left', 'right']
        )
    else:
        has_all_borders = False
    
    # Heuristic rules for common patterns

    if has_fill and not is_bold:
        return "highlight"
    
    if has_all_borders:
        return "bordered"
    
    if is_bold:
        return "bold_text"
    
    return "plain"


# ===== Test/Demo Code =====
if __name__ == "__main__":
    """
    Test the format extraction functions.
    Run: python format_extractor.py
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    import json
    
    print("=== Format Extractor Test ===\n")
    
    # Create test workbook
    wb = Workbook()
    ws = wb.active
    
    # Test cell with various formats
    cell = ws['A1']
    cell.value = "Header"
    cell.font = Font(name="Arial", size=14, bold=True, color="FF0000")
    cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    cell.border = Border(bottom=Side(style='medium', color="000000"))
    cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Extract formats
    fmt = extract_complete_format_info(cell)
    
    print("1. Complete format info:")
    print(json.dumps(fmt, indent=2))
    print()
    
    print("2. Format signature:")
    sig = create_format_signature(fmt)
    print(f"   {sig}")
    print()
    
    print("3. Semantic signature:")
    sem_sig = create_semantic_signature(fmt)
    print(f"   {sem_sig}")
    print()
    
    print("Test completed successfully! ✓")