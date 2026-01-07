"""
Spreadsheet Encoder Module - SpreadsheetLLM-style encoding utilities.

This module implements a lightweight version of SpreadsheetLLM-like compression:
- Finds "structural anchors" (important rows/columns) using Connected Components algorithm
- Keeps a neighborhood around anchors
- Builds an inverted index (value -> cell ranges) and format regions
- Optionally detects multiple sub-tables separated by blank rows/cols

IMPROVED: Replaced O(R⁴C⁴) heuristic anchor detection with O(RC) Connected Components algorithm
"""

from __future__ import annotations

import json
import logging
import os
import re
import tempfile
from collections import defaultdict
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Sequence, Tuple, Set
from itertools import islice  # ✅ FIX #2: Import islice

import numpy as np
import openpyxl
import pandas as pd
from openpyxl.utils import column_index_from_string, get_column_letter

logger = logging.getLogger(__name__)

EMAIL_REGEX = re.compile(r"^[a-zA-Z0-9_.+-]+@[a-zA-Z0-9-]+\.[a-zA-Z0-9-.]+$")


# ==============================================================================
# SpreadsheetLLM Helper Functions (UNCHANGED)
# ==============================================================================

def infer_cell_data_type(cell) -> str:
    """
    Infer a fine-grained cell type using openpyxl cell metadata + value.
    Returns: empty | email | text | numeric | boolean | datetime | error | formula* | unknown
    """
    if cell.value is None:
        return "empty"

    if isinstance(cell.value, str) and EMAIL_REGEX.match(cell.value):
        return "email"

    data_type = getattr(cell, "data_type", None)

    if data_type == "s":
        return "text"
    if data_type == "n":
        return "numeric"
    if data_type == "b":
        return "boolean"
    if data_type == "d":
        return "datetime"
    if data_type == "e":
        return "error"

    # Formula handling (depends on whether workbook is loaded with data_only)
    if data_type == "f":
        if cell.value is not None:
            if isinstance(cell.value, str):
                return "text"
            if isinstance(cell.value, (int, float)):
                return "numeric"
            if isinstance(cell.value, bool):
                return "boolean"
            return "formula_cached_value"
        return "formula"

    # "g" is used for rich text in some cases
    if data_type == "g":
        return "text" if cell.value is not None else "empty"

    return "unknown"


def get_number_format_string(cell) -> str:
    try:
        nfs = cell.number_format
        if nfs is None or nfs == "":
            return "General"
        return str(nfs)
    except Exception:
        return "General"


def categorize_number_format(number_format_string: str, cell) -> str:
    """
    Categorize Excel number format string into semantic buckets.
    Only relevant if inferred type is numeric/datetime.
    """
    cell_data_type = infer_cell_data_type(cell)
    if cell_data_type not in {"numeric", "datetime"}:
        return "not_applicable"

    if number_format_string is None or str(number_format_string).lower() == "general":
        return "datetime_general" if cell_data_type == "datetime" else "general"

    if number_format_string == "@" or str(number_format_string).lower() == "text":
        return "text_format"

    if any(c in str(number_format_string) for c in ["$", "€", "£", "¥"]):
        return "currency"

    if "%" in str(number_format_string):
        return "percentage"

    nf_lower = str(number_format_string).lower()
    if "e+" in nf_lower or "e-" in nf_lower:
        return "scientific"

    if "#" in nf_lower and "/" in nf_lower and "?" in nf_lower:
        return "fraction"

    date_keywords = ["yyyy", "yy", "mmmm", "mmm", "mm", "dddd", "ddd", "dd", "d"]
    time_keywords = ["hh", "h", "ss", "s", "am/pm", "a/p"]

    is_date = any(k in nf_lower for k in date_keywords)
    is_time = any(k in nf_lower for k in time_keywords)

    if ":" in str(number_format_string):
        tmp = str(number_format_string).replace("0", "").replace("#", "").replace(",", "").replace(".", "")
        if ":" in tmp:
            is_time = True

    if is_date and is_time:
        return "datetime_custom"
    if is_date:
        return "date_custom"
    if is_time:
        return "time_custom"

    if cell_data_type == "numeric":
        if number_format_string in {"0", "#,##0"}:
            return "integer"
        if number_format_string in {"0.00", "#,##0.00", "0.0", "#,##0.0"}:
            return "float"
        return "other_numeric"

    if cell_data_type == "datetime":
        return "other_date"

    return "unknown_format_category"


def detect_semantic_type(cell) -> str:
    """
    Detect semantic type combining inferred cell data type and format patterns.
    """
    data_type = infer_cell_data_type(cell)
    if data_type == "email":
        return "email"

    nfs = get_number_format_string(cell)
    category = categorize_number_format(nfs, cell)
    nfs_lower = nfs.lower()

    if category == "percentage":
        return "percentage"
    if category == "currency":
        return "currency"
    if category in {"date_custom", "datetime_custom", "datetime_general", "other_date"}:
        if ("yyyy" in nfs_lower or "yy" in nfs_lower) and not any(x in nfs_lower for x in ["m", "d"]):
            return "year"
        return "date"
    if category == "time_custom":
        return "time"
    if category == "scientific":
        return "scientific_notation"

    if data_type == "numeric":
        if isinstance(cell.value, int) or category == "integer":
            return "integer"
        if isinstance(cell.value, float) or category == "float":
            return "float"
        return "numeric"

    return data_type


def split_cell_ref(cell_ref: str) -> Tuple[str, int]:
    col_str = "".join(filter(str.isalpha, cell_ref))
    row_str = "".join(filter(str.isdigit, cell_ref))
    return col_str, int(row_str) if row_str else 0


def _merge_refs(refs: Sequence[str]) -> List[str]:
    """
    Merge a list of single-cell refs (e.g., A1, B1, A2) into a list of
    compact refs/ranges (e.g., A1:C1, A2, ...).
    """
    result: List[str] = []
    singles: List[str] = []
    for ref in refs:
        if ":" in str(ref):
            result.append(str(ref))
        else:
            singles.append(str(ref))

    coords: List[Tuple[int, int]] = []
    for ref in sorted(set(singles)):
        try:
            col_letter, row = split_cell_ref(ref)
            col = column_index_from_string(col_letter)
            coords.append((row, col))
        except Exception:
            continue

    cell_set = set(coords)
    processed = set()
    ranges: List[str] = []

    for row, col in sorted(coords):
        if (row, col) in processed:
            continue

        width = 1
        while (row, col + width) in cell_set and (row, col + width) not in processed:
            width += 1

        height = 1
        expanding = True
        while expanding:
            next_row = row + height
            for w in range(width):
                if (next_row, col + w) not in cell_set or (next_row, col + w) in processed:
                    expanding = False
                    break
            if expanding:
                height += 1

        end_col = col + width - 1
        end_row = row + height - 1
        start_ref = f"{get_column_letter(col)}{row}"
        end_ref = f"{get_column_letter(end_col)}{end_row}"

        if width == 1 and height == 1:
            ranges.append(start_ref)
        else:
            ranges.append(f"{start_ref}:{end_ref}")

        for r in range(row, row + height):
            for c in range(col, col + width):
                processed.add((r, c))

    return result + ranges


@dataclass(frozen=True)
class SheetRegion:
    min_row: int
    max_row: int
    min_col: int
    max_col: int

    def area(self) -> int:
        return max(0, self.max_row - self.min_row + 1) * max(0, self.max_col - self.min_col + 1)

    def to_dict(self) -> Dict[str, Any]:
        return {
            "min_row": self.min_row,
            "max_row": self.max_row,
            "min_col": self.min_col,
            "max_col": self.max_col,
        }


def _iter_values_only(sheet, region: SheetRegion) -> Iterable[Tuple[int, int, Any]]:
    for r_idx, row in enumerate(
            sheet.iter_rows(
                min_row=region.min_row,
                max_row=region.max_row,
                min_col=region.min_col,
                max_col=region.max_col,
                values_only=True,
            ),
            start=region.min_row,
    ):
        for c_offset, v in enumerate(row, start=region.min_col):
            yield r_idx, c_offset, v


def detect_table_regions(
        sheet,
        *,
        full_region: Optional[SheetRegion] = None,
        min_nonempty_cells: int = 8,
) -> List[SheetRegion]:
    """
    Detect sub-table blocks separated by fully-empty rows and/or columns.
    Works best for spreadsheets with blank lines between tables.
    """
    region = full_region or SheetRegion(1, sheet.max_row or 1, 1, sheet.max_column or 1)

    row_has = defaultdict(int)
    col_has = defaultdict(int)

    for r, c, v in _iter_values_only(sheet, region):
        if v is None or (isinstance(v, str) and v.strip() == ""):
            continue
        row_has[r] += 1
        col_has[c] += 1

    nonempty_rows = sorted(row_has.keys())
    nonempty_cols = sorted(col_has.keys())

    if not nonempty_rows or not nonempty_cols:
        return []

    row_segments: List[Tuple[int, int]] = []
    start = prev = nonempty_rows[0]
    for r in nonempty_rows[1:]:
        if r == prev + 1:
            prev = r
            continue
        row_segments.append((start, prev))
        start = prev = r
    row_segments.append((start, prev))

    col_segments: List[Tuple[int, int]] = []
    start = prev = nonempty_cols[0]
    for c in nonempty_cols[1:]:
        if c == prev + 1:
            prev = c
            continue
        col_segments.append((start, prev))
        start = prev = c
    col_segments.append((start, prev))

    regions: List[SheetRegion] = []
    for (r0, r1) in row_segments:
        for (c0, c1) in col_segments:
            cand = SheetRegion(r0, r1, c0, c1)
            nonempty = 0
            for _, _, v in _iter_values_only(sheet, cand):
                if v is None or (isinstance(v, str) and v.strip() == ""):
                    continue
                nonempty += 1
                if nonempty >= min_nonempty_cells:
                    break
            if nonempty >= min_nonempty_cells:
                regions.append(cand)

    if not regions:
        regions.append(SheetRegion(min(nonempty_rows), max(nonempty_rows), min(nonempty_cols), max(nonempty_cols)))

    regions.sort(key=lambda r: (r.area(), r.min_row, r.min_col), reverse=True)
    return regions


# ==============================================================================
# IMPROVED: Connected Components Anchor Detection (O(RC) complexity)
# ==============================================================================

def get_fill_color(cell) -> str:
    """Extract fill color from a cell for style comparison."""
    try:
        fill = cell.fill
        if fill and fill.patternType == 'solid':
            if fill.fgColor and fill.fgColor.rgb:
                return str(fill.fgColor.rgb)
            if fill.start_color and fill.start_color.rgb:
                return str(fill.start_color.rgb)
            if fill.fgColor and fill.fgColor.index:
                return f"index_{fill.fgColor.index}"
    except Exception:
        pass
    return 'none'


def has_border(cell) -> bool:
    """Check if a cell has any border styling."""
    try:
        border = cell.border
        return any([
            border.left and border.left.style,
            border.right and border.right.style,
            border.top and border.top.style,
            border.bottom and border.bottom.style
        ])
    except Exception:
        return False


def cells_similar(cell1, cell2, threshold: float = 0.75) -> bool:
    """
    Determine if two cells should belong to the same connected region.
    Uses weighted scoring: bold(2.0) + fill(3.0) + border(1.5) + type(2.0) + format(1.5)
    """
    score = 0.0
    total = 0.0

    # Font bold (weight: 2.0)
    try:
        bold1 = cell1.font.bold if cell1.font else False
        bold2 = cell2.font.bold if cell2.font else False
        if bold1 == bold2:
            score += 2.0
    except Exception:
        pass
    total += 2.0

    # Fill color (weight: 3.0)
    if get_fill_color(cell1) == get_fill_color(cell2):
        score += 3.0
    total += 3.0

    # Border (weight: 1.5)
    if has_border(cell1) == has_border(cell2):
        score += 1.5
    total += 1.5

    # Data type (weight: 2.0)
    type1 = type(cell1.value) if cell1.value is not None else type(None)
    type2 = type(cell2.value) if cell2.value is not None else type(None)
    if type1 == type2:
        score += 2.0
    total += 2.0

    # Number format (weight: 1.5)
    try:
        fmt1 = cell1.number_format if cell1.number_format else 'General'
        fmt2 = cell2.number_format if cell2.number_format else 'General'
        if fmt1 == fmt2:
            score += 1.5
    except Exception:
        pass
    total += 1.5

    return (score / total) >= threshold


def build_connectivity_graph(sheet, region: Optional[SheetRegion] = None) -> Dict[str, Set[str]]:
    """
    Build undirected graph of similar adjacent cells. O(RC) complexity.
    Each cell is visited once, checking right and bottom neighbors only.
    """
    if region:
        min_r, max_r = region.min_row, region.max_row
        min_c, max_c = region.min_col, region.max_col
    else:
        min_r, max_r = 1, sheet.max_row
        min_c, max_c = 1, sheet.max_column

    graph = defaultdict(set)

    for r in range(min_r, max_r + 1):
        for c in range(min_c, max_c + 1):
            cell = sheet.cell(row=r, column=c)

            # Skip empty cells
            if cell.value is None or (isinstance(cell.value, str) and cell.value.strip() == ''):
                continue

            cell_ref = f"{get_column_letter(c)}{r}"

            # Check right neighbor
            if c < max_c:
                right_cell = sheet.cell(row=r, column=c + 1)
                right_val = right_cell.value
                if right_val is not None and \
                        not (isinstance(right_val, str) and right_val.strip() == '') and \
                        cells_similar(cell, right_cell):
                    right_ref = f"{get_column_letter(c + 1)}{r}"
                    graph[cell_ref].add(right_ref)
                    graph[right_ref].add(cell_ref)

            # Check bottom neighbor
            if r < max_r:
                bottom_cell = sheet.cell(row=r + 1, column=c)
                bottom_val = bottom_cell.value
                if bottom_val is not None and \
                        not (isinstance(bottom_val, str) and bottom_val.strip() == '') and \
                        cells_similar(cell, bottom_cell):
                    bottom_ref = f"{get_column_letter(c)}{r + 1}"
                    graph[cell_ref].add(bottom_ref)
                    graph[bottom_ref].add(cell_ref)

    return dict(graph)


def find_connected_components(graph: Dict[str, Set[str]],
                              min_component_size: int = 4) -> List[List[str]]:
    """
    Find all connected components via DFS. O(V+E) = O(RC) complexity.
    Filters out components smaller than min_component_size to reduce noise.
    """
    visited = set()
    components = []

    for start_ref in graph:
        if start_ref in visited:
            continue

        # DFS traversal
        component = []
        stack = [start_ref]

        while stack:
            ref = stack.pop()
            if ref in visited:
                continue

            visited.add(ref)
            component.append(ref)

            for neighbor in graph.get(ref, []):
                if neighbor not in visited:
                    stack.append(neighbor)

        if len(component) >= min_component_size:
            components.append(component)

    # Sort by size (largest first)
    components.sort(key=len, reverse=True)
    return components


def extract_component_boundaries(components: List[List[str]]) -> Tuple[Set[int], Set[int]]:
    """
    Extract bounding box boundaries from each component as anchors. O(total_cells).
    Returns row and column boundary sets.
    """
    row_boundaries = set()
    col_boundaries = set()

    for component in components:
        if not component:
            continue

        min_r = min_c = float('inf')
        max_r = max_c = 0

        for cell_ref in component:
            col_letter, row = split_cell_ref(cell_ref)
            col = column_index_from_string(col_letter)

            min_r = min(min_r, row)
            max_r = max(max_r, row)
            min_c = min(min_c, col)
            max_c = max(max_c, col)

        # Add all four boundaries
        row_boundaries.add(min_r)
        row_boundaries.add(max_r)
        col_boundaries.add(min_c)
        col_boundaries.add(max_c)

    return row_boundaries, col_boundaries


def expand_with_k_neighborhood(indices: Set[int], k: int, max_index: int) -> List[int]:
    """
    Expand anchor indices with k-neighborhood. O(|indices| * k).
    Ensures all indices are within valid range [1, max_index].
    """
    expanded = set()

    for idx in indices:
        for offset in range(-k, k + 1):
            expanded_idx = idx + offset
            if 1 <= expanded_idx <= max_index:
                expanded.add(expanded_idx)

    return sorted(expanded)


def fast_find_structural_anchors(sheet, k: int, *, region: Optional[SheetRegion] = None) -> Tuple[List[int], List[int]]:
    """
    Find structural anchors using Connected Components algorithm. O(RC) complexity.

    This replaces the original CV-based method with a more robust algorithm that:
    1. Builds connectivity graph of similarly-styled cells - O(RC)
    2. Finds connected components via DFS - O(RC)
    3. Extracts component boundaries as anchors - O(RC)
    4. Expands with k-neighborhood - O(anchors * k)

    Fallback: If no components found, uses full sheet boundaries.
    """
    reg = region or SheetRegion(1, sheet.max_row or 1, 1, sheet.max_column or 1)

    logger.debug(f"Finding structural anchors with Connected Components (k={k})")

    # Phase 1: Build graph - O(RC)
    graph = build_connectivity_graph(sheet, region)

    if not graph or len(graph) < 4:
        logger.warning("Insufficient cells for component analysis, using boundaries")
        return ([reg.min_row, reg.max_row], [reg.min_col, reg.max_col])

    # Phase 2: Find components - O(RC)
    components = find_connected_components(graph, min_component_size=4)

    if not components:
        logger.warning("No components found, using boundaries")
        return ([reg.min_row, reg.max_row], [reg.min_col, reg.max_col])

    # Phase 3: Extract boundaries - O(total_cells)
    row_boundaries, col_boundaries = extract_component_boundaries(components)

    # Phase 4: Expand with k-neighborhood - O(anchors * k)
    row_anchors = expand_with_k_neighborhood(row_boundaries, k, reg.max_row)
    col_anchors = expand_with_k_neighborhood(col_boundaries, k, reg.max_col)

    # Ensure anchors are within region bounds
    row_anchors = [r for r in row_anchors if reg.min_row <= r <= reg.max_row]
    col_anchors = [c for c in col_anchors if reg.min_col <= c <= reg.max_col]

    logger.debug(f"Found {len(row_anchors)} row anchors, {len(col_anchors)} col anchors "
                 f"from {len(components)} components")

    return sorted(set(row_anchors)), sorted(set(col_anchors))


# ==============================================================================
# Rest of encoding pipeline (UNCHANGED)
# ==============================================================================

def extract_cells_near_anchors(sheet, row_anchors: List[int], col_anchors: List[int], k: int, *, region: Optional[SheetRegion] = None) -> Tuple[List[int], List[int]]:
    """Keep k-neighborhood rows/cols around each anchor within region bounds."""
    reg = region or SheetRegion(1, sheet.max_row or 1, 1, sheet.max_column or 1)

    rows_to_keep = set()
    cols_to_keep = set()

    for r in row_anchors:
        for rr in range(max(reg.min_row, r - k), min(reg.max_row, r + k) + 1):
            rows_to_keep.add(rr)

    for c in col_anchors:
        for cc in range(max(reg.min_col, c - k), min(reg.max_col, c + k) + 1):
            cols_to_keep.add(cc)

    return sorted(rows_to_keep), sorted(cols_to_keep)


def compress_homogeneous_regions(sheet, kept_rows: List[int], kept_cols: List[int], *, region: Optional[SheetRegion] = None) -> Tuple[List[int], List[int]]:
    """Remove entirely empty rows/cols from kept grid. Keeps at least one row/col."""
    reg = region or SheetRegion(1, sheet.max_row or 1, 1, sheet.max_column or 1)
    if not kept_rows or not kept_cols:
        return kept_rows, kept_cols

    def row_nonempty(r: int) -> bool:
        for c in kept_cols:
            if not (reg.min_col <= c <= reg.max_col):
                continue
            v = sheet.cell(row=r, column=c).value
            if v is None:
                continue
            if isinstance(v, str) and v.strip() == "":
                continue
            return True
        return False

    def col_nonempty(c: int) -> bool:
        for r in kept_rows:
            if not (reg.min_row <= r <= reg.max_row):
                continue
            v = sheet.cell(row=r, column=c).value
            if v is None:
                continue
            if isinstance(v, str) and v.strip() == "":
                continue
            return True
        return False

    kept_rows2 = [r for r in kept_rows if row_nonempty(r)]
    kept_cols2 = [c for c in kept_cols if col_nonempty(c)]

    if not kept_rows2:
        kept_rows2 = kept_rows[:1]
    if not kept_cols2:
        kept_cols2 = kept_cols[:1]

    return kept_rows2, kept_cols2


def _merged_start_cells(sheet) -> Dict[str, str]:
    """Build mapping: start_cell.coordinate -> range string (e.g., "A1:C3")"""
    start_map: Dict[str, str] = {}
    try:
        for m_range in sheet.merged_cells.ranges:
            start_map[m_range.start_cell.coordinate] = str(m_range)
    except Exception:
        pass
    return start_map


def _stringify_cell_value(v: Any, *, max_len: int = 200) -> str:
    if v is None:
        return ""
    try:
        s = str(v)
    except Exception:
        s = "ERROR_VALUE"
    if len(s) > max_len:
        s = s[:max_len] + "…"
    return s


def create_inverted_index(
        sheet,
        kept_rows: List[int],
        kept_cols: List[int],
        *,
        max_value_len: int = 200,
) -> Tuple[Dict[str, List[str]], Dict[str, List[str]]]:
    """
    Create inverted index and format map from kept cells.
    """
    inverted_index: Dict[str, List[str]] = defaultdict(list)
    format_map: Dict[str, List[str]] = defaultdict(list)

    merged_start_map = _merged_start_cells(sheet)

    for r in kept_rows:
        for c in kept_cols:
            cell = sheet.cell(row=r, column=c)
            cell_ref = f"{get_column_letter(c)}{r}"

            merged_range_str = merged_start_map.get(cell.coordinate)

            # Value index
            try:
                if merged_range_str is not None:
                    v = cell.value
                    if v is not None and not (isinstance(v, str) and v.strip() == ""):
                        v_str = _stringify_cell_value(v, max_len=max_value_len)
                        inverted_index[v_str].append(merged_range_str)
                else:
                    v = cell.value
                    if v is not None and not (isinstance(v, str) and v.strip() == ""):
                        v_str = _stringify_cell_value(v, max_len=max_value_len)
                        inverted_index[v_str].append(cell_ref)
            except Exception:
                inverted_index["ERROR_VALUE"].append(cell_ref)

            # Format map
            try:
                format_info: Dict[str, Any] = {}
                format_info["font"] = {"bold": getattr(cell.font, "bold", None)}
                alignment = getattr(cell, "alignment", None)
                format_info["alignment"] = {"horizontal": getattr(alignment, "horizontal", None)}

                original_number_format = get_number_format_string(cell)
                inferred_type = infer_cell_data_type(cell)
                category = categorize_number_format(original_number_format, cell)

                format_info["original_number_format"] = original_number_format
                format_info["inferred_data_type"] = inferred_type
                format_info["number_format_category"] = category
                format_info["merged"] = merged_range_str is not None
                if merged_range_str is not None:
                    format_info["merged_range"] = merged_range_str

                format_key = json.dumps(format_info, sort_keys=True)
                format_map[format_key].append(cell_ref)
            except Exception:
                pass

    return dict(inverted_index), dict(format_map)


def create_inverted_index_translation(inverted_index: Dict[str, List[str]]) -> Dict[str, List[str]]:
    """Merge single-cell refs into ranges; keep pre-merged ranges intact."""
    merged_index: Dict[str, List[str]] = {}
    for value, refs in inverted_index.items():
        if value is None or str(value).strip() == "":
            continue
        merged_index[str(value)] = _merge_refs(refs)
    return merged_index


def aggregate_formats(sheet, format_map: Dict[str, List[str]]) -> Dict[str, List[str]]:
    """Aggregate formats by semantic-type + number-format, merging into rectangles."""
    aggregated_formats: Dict[str, List[str]] = defaultdict(list)
    processed_cells = set()

    type_nfs_map: Dict[str, List[str]] = defaultdict(list)
    for _, cells in format_map.items():
        for cell_ref in cells:
            try:
                cell = sheet[cell_ref]
            except Exception:
                continue
            nfs = get_number_format_string(cell)
            sem_type = detect_semantic_type(cell)
            key = json.dumps({"type": sem_type, "nfs": nfs}, sort_keys=True)
            type_nfs_map[key].append(cell_ref)

    for key, cells in type_nfs_map.items():
        cells_set = set(cells)
        for start_cell in cells:
            if start_cell in processed_cells:
                continue

            try:
                start_col_letter, start_row = split_cell_ref(start_cell)
                start_col = column_index_from_string(start_col_letter)
            except Exception:
                continue

            best_width = 1
            best_height = 1
            best_area = 1
            best_end_cell = start_cell

            max_width = min(20, sheet.max_column - start_col + 1)
            max_height = min(20, sheet.max_row - start_row + 1)

            for width in range(1, max_width + 1):
                for height in range(1, max_height + 1):
                    valid_rectangle = True
                    for r in range(start_row, start_row + height):
                        for c in range(start_col, start_col + width):
                            cell_ref = f"{get_column_letter(c)}{r}"
                            if cell_ref not in cells_set or cell_ref in processed_cells:
                                valid_rectangle = False
                                break
                        if not valid_rectangle:
                            break

                    if valid_rectangle:
                        area = width * height
                        if area > best_area:
                            best_width = width
                            best_height = height
                            best_area = area
                            best_end_cell = f"{get_column_letter(start_col + width - 1)}{start_row + height - 1}"

            region = start_cell if best_width == 1 and best_height == 1 else f"{start_cell}:{best_end_cell}"
            aggregated_formats[key].append(region)

            for r in range(start_row, start_row + best_height):
                for c in range(start_col, start_col + best_width):
                    processed_cells.add(f"{get_column_letter(c)}{r}")

    return dict(aggregated_formats)


def cluster_numeric_ranges(sheet, format_map: Dict[str, List[str]]) -> Dict[str, List[str]]:
    """Find numeric regions by grouping cells with numeric type and similar format."""
    numeric_map = {fmt: cells for fmt, cells in format_map.items() if json.loads(fmt).get("inferred_data_type") == "numeric"}

    if not numeric_map:
        for r in range(1, sheet.max_row + 1):
            for c in range(1, sheet.max_column + 1):
                cell = sheet.cell(row=r, column=c)
                if infer_cell_data_type(cell) == "numeric":
                    fmt_key = json.dumps({"type": detect_semantic_type(cell), "nfs": get_number_format_string(cell)}, sort_keys=True)
                    numeric_map.setdefault(fmt_key, []).append(f"{get_column_letter(c)}{r}")

    return aggregate_formats(sheet, numeric_map)


def spreadsheet_llm_encode_with_helpers(
        excel_path: str,
        output_path: Optional[str] = None,
        k: int = 2,
        *,
        data_only: bool = True,
        detect_subtables: bool = True,
        max_value_len: int = 200,
        min_nonempty_cells_for_region: int = 8,
) -> Dict[str, Any]:
    """Encode Excel file into compact JSON structure using SpreadsheetLLM methodology."""
    wb = openpyxl.load_workbook(excel_path, data_only=data_only)

    sheets_encoding: Dict[str, Any] = {}
    compression_metrics: Dict[str, Any] = {"sheets": {}}
    overall_orig = overall_anchor = overall_index = overall_format = overall_final = 0

    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        if (sheet.max_row or 0) <= 1 and (sheet.max_column or 0) <= 1:
            continue

        full_region = SheetRegion(1, sheet.max_row or 1, 1, sheet.max_column or 1)

        regions: List[SheetRegion]
        if detect_subtables:
            regions = detect_table_regions(sheet, full_region=full_region, min_nonempty_cells=min_nonempty_cells_for_region)
        else:
            regions = [full_region]

        region_encodings: List[Dict[str, Any]] = []

        # ✅ FIX #3: Calculate original_tokens once, outside region loop
        original_cells = {}
        for r, c, v in _iter_values_only(sheet, full_region):
            if v is None:
                continue
            s = _stringify_cell_value(v, max_len=max_value_len)
            if s.strip() == "":
                continue
            original_cells[f"{get_column_letter(c)}{r}"] = s
        original_tokens = len(json.dumps(original_cells, ensure_ascii=False))

        for reg_idx, reg in enumerate(regions):
            row_anchors, col_anchors = fast_find_structural_anchors(sheet, k, region=reg)

            kept_rows, kept_cols = extract_cells_near_anchors(sheet, row_anchors, col_anchors, k, region=reg)

            if not kept_rows or not kept_cols:
                kept_rows = list(range(reg.min_row, min(reg.min_row + 20, reg.max_row + 1)))
                kept_cols = list(range(reg.min_col, min(reg.min_col + 20, reg.max_col + 1)))

            kept_rows, kept_cols = compress_homogeneous_regions(sheet, kept_rows, kept_cols, region=reg)

            anchor_cells = {}
            for r in kept_rows:
                for c in kept_cols:
                    v = sheet.cell(row=r, column=c).value
                    if v is None:
                        continue
                    s = _stringify_cell_value(v, max_len=max_value_len)
                    if s.strip() == "":
                        continue
                    anchor_cells[f"{get_column_letter(c)}{r}"] = s
            anchor_tokens = len(json.dumps(anchor_cells, ensure_ascii=False))

            inverted_index, format_map = create_inverted_index(sheet, kept_rows, kept_cols, max_value_len=max_value_len)
            index_tokens = len(json.dumps(inverted_index, ensure_ascii=False))

            merged_index = create_inverted_index_translation(inverted_index)
            merged_tokens = len(json.dumps(merged_index, ensure_ascii=False))

            aggregated_formats = aggregate_formats(sheet, format_map)
            format_tokens = len(json.dumps(aggregated_formats, ensure_ascii=False))

            numeric_ranges = cluster_numeric_ranges(sheet, format_map)
            numeric_tokens = len(json.dumps(numeric_ranges, ensure_ascii=False))

            sheet_encoding = {
                "region": reg.to_dict(),
                "structural_anchors": {
                    "rows": row_anchors,
                    "columns": [get_column_letter(c) for c in col_anchors],
                },
                "cells": merged_index,
                "formats": aggregated_formats,
                "numeric_ranges": numeric_ranges,
            }
            final_tokens = len(json.dumps(sheet_encoding, ensure_ascii=False))

            region_encodings.append(
                {
                    "region_index": reg_idx,
                    "encoding": sheet_encoding,
                    "metrics": {
                        "original_tokens": original_tokens,  # ✅ Use pre-calculated value
                        "after_anchor_tokens": anchor_tokens,
                        "after_inverted_index_tokens": index_tokens,
                        "after_merged_index_tokens": merged_tokens,
                        "after_format_tokens": format_tokens,
                        "numeric_tokens": numeric_tokens,
                        "final_tokens": final_tokens,
                        "anchor_ratio": (original_tokens / anchor_tokens) if anchor_tokens else 0,
                        "index_ratio": (original_tokens / index_tokens) if index_tokens else 0,
                        "format_ratio": (original_tokens / format_tokens) if format_tokens else 0,
                        "overall_ratio": (original_tokens / final_tokens) if final_tokens else 0,
                    },
                }
            )

        primary = region_encodings[0]["encoding"] if region_encodings else {}
        primary_metrics = region_encodings[0]["metrics"] if region_encodings else {
            "original_tokens": original_tokens,
            "final_tokens": 0,
            "overall_ratio": 0,
        }

        sheets_encoding[sheet_name] = {
            **primary,
            "regions": [r["encoding"] for r in region_encodings],
        }

        compression_metrics["sheets"][sheet_name] = {
            **primary_metrics,
            "num_regions": len(region_encodings),
        }

        overall_orig += original_tokens
        overall_anchor += int(primary_metrics.get("after_anchor_tokens", 0) or 0)
        overall_index += int(primary_metrics.get("after_inverted_index_tokens", 0) or 0)
        overall_format += int(primary_metrics.get("after_format_tokens", 0) or 0)
        overall_final += int(primary_metrics.get("final_tokens", 0) or 0)

    compression_metrics["overall"] = {
        "original_tokens": overall_orig,
        "after_anchor_tokens": overall_anchor,
        "after_inverted_index_tokens": overall_index,
        "after_format_tokens": overall_format,
        "final_tokens": overall_final,
        "overall_ratio": (overall_orig / overall_final) if overall_final else 0,
    }

    full_encoding = {
        "file_name": os.path.basename(excel_path),
        "sheets": sheets_encoding,
        "compression_metrics": compression_metrics,
    }

    if output_path:
        with open(output_path, "w", encoding="utf-8") as f:
            json.dump(full_encoding, f, ensure_ascii=False, indent=2)
        logger.info("Saved encoding to %s", output_path)

    return full_encoding


def convert_csv_to_xlsx(csv_path: str, xlsx_path: str) -> bool:
    """Best-effort CSV reader: tries multiple encodings and separators."""
    tried: List[Tuple[str, str, str]] = []

    for enc in ["utf-16", "utf-8", "latin1"]:
        for sep in ["\t", ",", ";", "|"]:
            try:
                df = pd.read_csv(csv_path, encoding=enc, sep=sep)
                if df.shape[1] > 1:
                    df.to_excel(xlsx_path, index=False, engine="openpyxl")
                    logger.info("Converted CSV -> XLSX (encoding=%s sep=%s)", enc, sep)
                    return True
            except Exception as e:
                tried.append((enc, sep, str(e)))
                continue

    df = pd.read_csv(csv_path, encoding="utf-8", sep=None, engine="python")
    df.to_excel(xlsx_path, index=False, engine="openpyxl")
    logger.info("Converted CSV -> XLSX using python engine")
    return True


# ==============================================================================
# Main SpreadsheetEncoder Class
# ==============================================================================

class SpreadsheetEncoder:
    """
    SpreadsheetLLM-style encoder with Connected Components anchor detection.
    Returns encoded_text, metadata, and spreadsheet_llm_encoding.
    """

    def __init__(self, config: Dict[str, Any]):
        self.config = config
        self.k = int(config.get("anchor_neighborhood", 2))
        self.detect_subtables = bool(config.get("detect_subtables", True))
        self.data_only = bool(config.get("data_only", True))
        self.max_value_len = int(config.get("max_value_len", 200))
        self.min_nonempty_cells_for_region = int(config.get("min_nonempty_cells_for_region", 8))
        self._working_file: Optional[str] = None
        self._temp_xlsx: Optional[str] = None

        logger.info(
            "Initialized SpreadsheetEncoder (Connected Components) k=%s detect_subtables=%s data_only=%s",
            self.k,
            self.detect_subtables,
            self.data_only,
        )

    def load_file(self, file_path: str) -> pd.DataFrame:
        """Load file and convert CSV if needed."""
        file_path_p = Path(file_path)
        if not file_path_p.exists():
            raise FileNotFoundError(f"File not found: {file_path_p}")

        logger.info("Loading file: %s", file_path_p)

        if file_path_p.suffix.lower() == ".csv":
            logger.info("Converting CSV to XLSX...")
            temp_xlsx = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
            temp_xlsx_path = temp_xlsx.name
            temp_xlsx.close()

            convert_csv_to_xlsx(str(file_path_p), temp_xlsx_path)
            self._temp_xlsx = temp_xlsx_path
            working_file = temp_xlsx_path
        else:
            working_file = str(file_path_p)
            self._temp_xlsx = None

        df = pd.read_excel(working_file, engine="openpyxl")
        logger.info("Loaded dataframe with shape %s", df.shape)

        self._working_file = working_file
        return df

    def encode(self, df: pd.DataFrame) -> Dict[str, Any]:
        """Encode using SpreadsheetLLM methodology with Connected Components anchors."""
        logger.info("Encoding with SpreadsheetLLM (Connected Components)...")

        try:
            # Create temporary file if needed
            if not self._working_file:
                temp_file = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
                temp_file_path = temp_file.name
                temp_file.close()
                df.to_excel(temp_file_path, index=False, engine="openpyxl")
                self._working_file = temp_file_path
                self._temp_xlsx = temp_file_path

            # Perform encoding
            full_encoding = spreadsheet_llm_encode_with_helpers(
                self._working_file,
                output_path=None,
                k=self.k,
                data_only=self.data_only,
                detect_subtables=self.detect_subtables,
                max_value_len=self.max_value_len,
                min_nonempty_cells_for_region=self.min_nonempty_cells_for_region,
            )

            sheet_name = list(full_encoding["sheets"].keys())[0] if full_encoding["sheets"] else None
            sheet_encoding = full_encoding["sheets"].get(sheet_name, {}) if sheet_name else {}

            encoded_text = self._create_readable_text(sheet_encoding, full_encoding["compression_metrics"])
            metadata = self._extract_metadata(df, sheet_encoding, full_encoding["compression_metrics"])

            result = {
                "encoded_text": encoded_text,
                "metadata": metadata,
                "original_shape": df.shape,
                "preview_shape": df.shape,
                "dataframe": df,
                "spreadsheet_llm_encoding": sheet_encoding,
                "compression_metrics": full_encoding["compression_metrics"],
            }

            logger.info("Encoding complete. Compression: %.2fx", metadata.get("compression_ratio", 0))
            return result

        finally:
            # ✅ FIX #1: Always cleanup temporary file, even if exception occurs
            if self._temp_xlsx and Path(self._temp_xlsx).exists():
                try:
                    Path(self._temp_xlsx).unlink()
                    logger.debug(f"Deleted temporary file: {self._temp_xlsx}")
                except Exception as e:
                    logger.warning(f"Failed to delete temporary file: {e}")

            # Reset state for next encode call
            self._working_file = None
            self._temp_xlsx = None

    def _create_readable_text(self, sheet_encoding: Dict[str, Any], compression_metrics: Dict[str, Any]) -> str:
        """Create LLM-friendly text representation."""
        lines: List[str] = []
        sheet_metrics = (
            list(compression_metrics.get("sheets", {}).values())[0]
            if compression_metrics.get("sheets")
            else {}
        )

        lines.append("=" * 80)
        lines.append("SPREADSHEET ENCODING (SpreadsheetLLM)")
        lines.append("=" * 80)
        lines.append(f"Compression Ratio: {sheet_metrics.get('overall_ratio', 0):.2f}x")
        lines.append("")

        if "region" in sheet_encoding:
            reg = sheet_encoding["region"]
            lines.append("## REGION")
            lines.append(f"Bounds: rows {reg.get('min_row')}..{reg.get('max_row')}, cols {reg.get('min_col')}..{reg.get('max_col')}")
            lines.append("")

        anchors = sheet_encoding.get("structural_anchors", {})
        lines.append("## STRUCTURAL ANCHORS")
        lines.append(f"Key Rows: {anchors.get('rows', [])}")
        lines.append(f"Key Columns: {anchors.get('columns', [])}")
        lines.append("")

        cells = sheet_encoding.get("cells", {})
        lines.append("## CELL VALUES")
        lines.append(f"Unique values: {len(cells)}")

        # ✅ FIX #2: Use islice instead of creating full list
        for value, ranges in islice(cells.items(), 20):
            ranges_str = ", ".join(ranges[:5])
            if len(ranges) > 5:
                ranges_str += f" (+{len(ranges) - 5} more)"
            v_show = value if len(str(value)) <= 80 else (str(value)[:80] + "…")
            lines.append(f"  '{v_show}': {ranges_str}")
        if len(cells) > 20:
            lines.append(f"  ... and {len(cells) - 20} more values")
        lines.append("")

        formats = sheet_encoding.get("formats", {})
        lines.append("## FORMAT REGIONS")
        lines.append(f"Format groups: {len(formats)}")

        # ✅ FIX #2: Use islice for formats too
        for fmt_key, ranges in islice(formats.items(), 10):
            try:
                fmt = json.loads(fmt_key)
                ranges_str = ", ".join(ranges[:3])
                if len(ranges) > 3:
                    ranges_str += f" (+{len(ranges) - 3})"
                lines.append(f"  Type: {fmt.get('type')}, Format: {fmt.get('nfs')}")
                lines.append(f"    Ranges: {ranges_str}")
            except Exception:
                continue
        lines.append("")
        return "\n".join(lines)

    def _extract_metadata(self, df: pd.DataFrame, sheet_encoding: Dict[str, Any], compression_metrics: Dict[str, Any]) -> Dict[str, Any]:
        """Extract comprehensive metadata."""
        sheet_metrics = (
            list(compression_metrics.get("sheets", {}).values())[0]
            if compression_metrics.get("sheets")
            else {}
        )

        metadata: Dict[str, Any] = {
            "num_rows": len(df),
            "num_cols": len(df.columns),
            "column_names": df.columns.tolist(),
            "dtypes": {col: str(dtype) for col, dtype in df.dtypes.items()},
            "null_counts": df.isnull().sum().to_dict(),
            "null_percentages": (df.isnull().sum() / len(df) * 100).to_dict() if len(df) else {},
            "compression_ratio": sheet_metrics.get("overall_ratio", 0),
            "compression_stages": {
                "original_tokens": sheet_metrics.get("original_tokens", 0),
                "after_anchors": sheet_metrics.get("after_anchor_tokens", 0),
                "final_tokens": sheet_metrics.get("final_tokens", 0),
            },
            "sample_values": {},
            "columns": {},
        }

        max_unique_values = int(self.config.get("max_unique_values", 100))
        sample_size = min(int(self.config.get("metadata_sample_size", 10000)), len(df))
        df_sample = df if len(df) <= sample_size else df.sample(sample_size, random_state=0)

        for col in df.columns:
            col_metadata = self._analyze_column(df[col], df_sample[col], max_unique_values)
            metadata["columns"][col] = col_metadata

            non_null = df[col].dropna()
            metadata["sample_values"][col] = [str(v)[:50] for v in non_null.head(3).tolist()] if len(non_null) else []

        issues: List[str] = []
        for col in df.columns:
            col_str = str(col)
            has_cn = any("\u4e00" <= c <= "\u9fff" for c in col_str)
            has_en = any(c.isalpha() and ord(c) < 128 for c in col_str)
            if has_cn and has_en:
                issues.append(f"Multi-language column: {col}")
        metadata["potential_issues"] = issues

        return metadata

    def _analyze_column(self, full_series: pd.Series, sample_series: pd.Series, max_unique_values: int) -> Dict[str, Any]:
        """Analyze column metadata."""
        col_metadata: Dict[str, Any] = {
            "dtype": str(full_series.dtype),
            "null_count": int(full_series.isnull().sum()),
            "null_percentage": float(full_series.isnull().sum() / len(full_series) * 100) if len(full_series) > 0 else 0.0,
            "unique_count": int(full_series.nunique(dropna=True)),
            "unique_values": [],
            "unique_values_truncated": False,
            "value_counts": {},
            "sample_values": [],
            "inferred_type": None,
            "potential_delimiters": [],
            "has_bilingual_content": False,
            "statistics": {},
        }

        unique_vals = full_series.dropna().unique()
        if len(unique_vals) <= max_unique_values:
            col_metadata["unique_values"] = [str(v) for v in unique_vals]
            col_metadata["unique_values_truncated"] = False
            vc = full_series.value_counts(dropna=True)
            col_metadata["value_counts"] = {str(k): int(v) for k, v in vc.head(50).items()}
        else:
            col_metadata["unique_values"] = [str(v) for v in unique_vals[:max_unique_values]]
            col_metadata["unique_values_truncated"] = True

        non_null_sample = sample_series.dropna().head(20)
        col_metadata["sample_values"] = [str(v) for v in non_null_sample]

        col_metadata["inferred_type"] = self._infer_column_type(full_series, col_metadata)

        if col_metadata["inferred_type"] == "string":
            col_metadata["potential_delimiters"] = self._detect_delimiters(col_metadata["sample_values"])

        col_metadata["has_bilingual_content"] = self._check_bilingual(col_metadata["sample_values"])

        if pd.api.types.is_numeric_dtype(full_series):
            s = full_series.dropna()
            if len(s):
                col_metadata["statistics"] = {
                    "min": float(s.min()),
                    "max": float(s.max()),
                    "mean": float(s.mean()),
                    "median": float(s.median()),
                    "std": float(s.std(ddof=0)) if len(s) > 1 else 0.0,
                }

        return col_metadata

    def _infer_column_type(self, series: pd.Series, col_metadata: Dict[str, Any]) -> str:
        """Infer semantic column type."""
        unique_count = int(col_metadata.get("unique_count", 0))
        total_count = int(len(series)) if len(series) else 0
        unique_values = col_metadata.get("unique_values", [])

        if pd.api.types.is_numeric_dtype(series):
            if pd.api.types.is_integer_dtype(series):
                s = series.dropna()
                if len(s):
                    min_val, max_val = int(s.min()), int(s.max())
                    if 1900 <= min_val <= 2100 and 1900 <= max_val <= 2100:
                        return "year"
                return "integer"
            return "numeric"

        if unique_count <= 10:
            unique_vals_lower = [str(v).lower().strip() for v in unique_values]
            boolean_patterns = [
                {"yes", "no"},
                {"y", "n"},
                {"true", "false"},
                {"t", "f"},
                {"1", "0"},
                {"是", "否"},
                {"有", "沒有", "无"},
                {"有", "没有"},
            ]
            na_indicators = {"n/a", "na", "not applicable", "不適用", "不适用", "none", "null", ""}
            has_na_value = any(val in na_indicators for val in unique_vals_lower)
            actual_vals = set(unique_vals_lower) - set(na_indicators)

            for pattern in boolean_patterns:
                if actual_vals <= pattern and len(actual_vals) >= 2:
                    return "boolean_with_na" if has_na_value else "boolean"

            if len(actual_vals) == 1 and has_na_value:
                single_val = list(actual_vals)[0]
                all_boolean_vals = set().union(*boolean_patterns)
                if single_val in all_boolean_vals:
                    return "boolean_with_na"

            return "categorical"

        if total_count and unique_count / total_count > 0.95:
            return "identifier"

        if total_count and unique_count / total_count < 0.5 and unique_count <= 50:
            return "categorical"

        if self._is_date_column(series):
            return "date"

        return "string"

    def _detect_delimiters(self, sample_values: List[str]) -> List[Dict[str, Any]]:
        if not sample_values or len(sample_values) < 3:
            return []

        delimiters_to_check = [
            (" - ", "space-dash-space"),
            ("-", "dash"),
            (" / ", "space-slash-space"),
            ("/", "slash"),
            (" | ", "space-pipe-space"),
            ("|", "pipe"),
            (":", "colon"),
            (";", "semicolon"),
            (",", "comma"),
            ("(", "open-paren"),
            (")", "close-paren"),
        ]

        detected: List[Dict[str, Any]] = []
        total_values = len(sample_values)

        for delimiter, name in delimiters_to_check:
            count = sum(1 for val in sample_values if delimiter in str(val))
            percentage = (count / total_values * 100) if total_values else 0.0
            if percentage >= 50:
                split_counts = [len(str(val).split(delimiter)) for val in sample_values if delimiter in str(val)]
                is_consistent = len(set(split_counts)) == 1 if split_counts else False
                num_parts = split_counts[0] if split_counts and is_consistent else None
                detected.append(
                    {
                        "delimiter": delimiter,
                        "name": name,
                        "frequency": count,
                        "percentage": percentage,
                        "is_consistent": is_consistent,
                        "num_parts": num_parts,
                        "sample_split": str(sample_values[0]).split(delimiter) if delimiter in str(sample_values[0]) else None,
                    }
                )

        detected.sort(key=lambda x: x["percentage"], reverse=True)
        return detected

    def _check_bilingual(self, sample_values: List[str]) -> bool:
        if not sample_values:
            return False
        for val in sample_values[:10]:
            s = str(val)
            has_chinese = any("\u4e00" <= c <= "\u9fff" for c in s)
            has_english = any(c.isalpha() and ord(c) < 128 for c in s)
            if has_chinese and has_english:
                return True
        return False

    def _is_date_column(self, series: pd.Series) -> bool:
        if pd.api.types.is_datetime64_any_dtype(series):
            return True

        sample = series.dropna().head(10)
        if len(sample) == 0:
            return False

        try:
            parsed = pd.to_datetime(sample, errors="coerce")
            valid_dates = int(parsed.notna().sum())
            return (valid_dates / len(sample)) > 0.7
        except Exception:
            return False