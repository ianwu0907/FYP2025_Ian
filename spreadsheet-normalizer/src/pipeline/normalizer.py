"""
Table Normalizer Pipeline (Enhanced)
Orchestrates the complete normalization process using semantic reasoning.

Pipeline stages:
1. Encoding: Compress spreadsheet to LLM-friendly representation
2. Structure Analysis: Semantic understanding of table structure
3. Schema Estimation: Derive ideal tidy schema
4. Transformation: Two-stage (Strategy → Code) generation and execution
"""

import logging
import pandas as pd
import numpy as np
import openpyxl
import tempfile
from typing import Dict, Any, Optional, List, Tuple
from pathlib import Path
from dataclasses import dataclass
import json
import time

from ..encoder import SpreadsheetEncoder
from ..analyzer import StructureAnalyzer
from ..estimator import SchemaEstimator
from ..generator import TransformationGenerator

logger = logging.getLogger(__name__)


@dataclass
class TableRegion:
    """Represents a detected table region within the sheet"""
    start_row: int
    end_row: int
    start_col: int
    end_col: int
    name: str
    confidence: float
    
    def to_dict(self) -> Dict[str, Any]:
        return {
            'start_row': self.start_row,
            'end_row': self.end_row,
            'start_col': self.start_col,
            'end_col': self.end_col,
            'name': self.name,
            'confidence': self.confidence,
            'shape': (self.end_row - self.start_row + 1, 
                     self.end_col - self.start_col + 1)
        }


class TableNormalizer:
    """
    Main pipeline that orchestrates the complete table normalization process.

    Uses LLM semantic reasoning throughout:
    - Structure analysis identifies patterns through understanding, not matching
    - Schema estimation derives ideal tidy format based on tidy data principles
    - Transformation uses free-form strategy generation
    """

    def __init__(self, config: Dict[str, Any]):
        """Initialize the normalizer pipeline."""
        self.config = config

        # Initialize all components
        logger.info("Initializing Table Normalizer Pipeline (Enhanced)...")
        logger.info("  Using LLM semantic reasoning for generalized normalization")

        self.encoder = SpreadsheetEncoder(config.get('encoder', {}))
        self.analyzer = StructureAnalyzer(config.get('analyzer', {}) | config.get('llm', {}))
        self.estimator = SchemaEstimator(config.get('estimator', {}) | config.get('llm', {}))
        self.generator = TransformationGenerator(config.get('generator', {}) | config.get('llm', {}))

        # Output configuration
        self.output_config = config.get('output', {})
        self.save_intermediate = config.get('logging', {}).get('save_intermediate', True)
        self.output_dir = Path(config.get('logging', {}).get('output_dir', 'output'))
        self.output_dir.mkdir(exist_ok=True)
        
        # Table splitting configuration
        splitting_config = config.get('table_splitting', {})
        self.enable_table_splitting = splitting_config.get('enabled', True)
        self.split_processing_mode = splitting_config.get('processing_mode', 'separate')
        
        # Detection mode: 'auto', 'hyperlink', or 'empty_row'
        self.detection_mode = splitting_config.get('detection_mode', 'auto')
        
        # Empty row detection parameters
        self.min_table_rows = splitting_config.get('min_table_rows', 3)
        self.min_table_cols = splitting_config.get('min_table_cols', 2)
        self.empty_threshold = splitting_config.get('empty_threshold', 0.8)
        self.min_gap_rows = splitting_config.get('min_gap_rows', 1)
        
        # Hyperlink detection parameters
        hyperlink_config = splitting_config.get('hyperlink_detection', {})
        self.main_sheet_name = hyperlink_config.get('main_sheet', None)
        self.follow_external = hyperlink_config.get('follow_external', False)

        logger.info("Pipeline initialized successfully")

    def normalize(self, input_file: str, output_file: Optional[str] = None) -> Dict[str, Any]:
        """
        Execute the complete normalization pipeline.

        Args:
            input_file: Path to input spreadsheet (.xlsx or .csv)
            output_file: Path to output file (default: output/normalized_output.csv)

        Returns:
            Dict containing:
            - normalized_df: The transformed DataFrame
            - output_path: Path to saved output
            - pipeline_log: Execution log
            - intermediate_results: All intermediate outputs
        """
        start_time = time.time()
        logger.info(f"Starting normalization pipeline for: {input_file}")

        pipeline_log = {
            'input_file': input_file,
            'start_time': time.strftime('%Y-%m-%d %H:%M:%S'),
            'stages': {}
        }

        try:
            # ================================================================
            # Stage 1: ENCODING
            # ================================================================
            logger.info("\n" + "=" * 80)
            logger.info("STAGE 1: SPREADSHEET ENCODING")
            logger.info("=" * 80)

            encoded_data = self._run_encoding(input_file)
            pipeline_log['stages']['encoding'] = {
                'status': 'success',
                'shape': encoded_data['original_shape'],
                'compression_ratio': encoded_data['metadata'].get('compression_ratio', 0)
            }

            if self.save_intermediate:
                self._save_intermediate('01_encoded_data.json', {
                    'metadata': encoded_data['metadata'],
                    'encoded_text_preview': encoded_data['encoded_text'][:2000]
                })

            # ================================================================
            # Stage 2: STRUCTURE ANALYSIS (LLM Semantic Reasoning)
            # ================================================================
            logger.info("\n" + "=" * 80)
            logger.info("STAGE 2: STRUCTURE ANALYSIS (Semantic Reasoning)")
            logger.info("=" * 80)

            structure_analysis = self._run_structure_analysis(encoded_data)
            pipeline_log['stages']['structure_analysis'] = {
                'status': 'success',
                'has_bilingual': structure_analysis.get('row_patterns', {}).get('has_bilingual_rows', False),
                'has_section_markers': structure_analysis.get('row_patterns', {}).get('has_section_markers', False),
                'has_implicit_aggregation': structure_analysis.get('implicit_aggregation', {}).get('has_implicit_aggregation', False),
                'transformation_complexity': structure_analysis.get('transformation_complexity', 'unknown')
            }

            if self.save_intermediate:
                self._save_intermediate('02_structure_analysis.json', structure_analysis)

            # ================================================================
            # Stage 3: SCHEMA ESTIMATION (Tidy Data Principles)
            # ================================================================
            logger.info("\n" + "=" * 80)
            logger.info("STAGE 3: SCHEMA ESTIMATION (Tidy Data Principles)")
            logger.info("=" * 80)

            schema = self._run_schema_estimation(encoded_data, structure_analysis)
            pipeline_log['stages']['schema_estimation'] = {
                'status': 'success',
                'observation_unit': schema.get('observation_unit', {}).get('description', 'unknown'),
                'num_target_columns': len(schema.get('target_columns', [])),
                'expected_rows': schema.get('expected_output', {}).get('row_count_estimate', 0)
            }

            if self.save_intermediate:
                self._save_intermediate('03_schema.json', schema)

            # ================================================================
            # Stage 4: TRANSFORMATION (Strategy → Code → Execute → Validate)
            # ================================================================
            logger.info("\n" + "=" * 80)
            logger.info("STAGE 4: TRANSFORMATION (Two-Stage Generation)")
            logger.info("=" * 80)

            transformation_result = self._run_transformation(
                encoded_data, structure_analysis, schema
            )
            pipeline_log['stages']['transformation'] = {
                'status': 'success',
                'output_shape': transformation_result['normalized_df'].shape,
                'validation': transformation_result['validation_result'],
                'attempts': transformation_result.get('attempts', 1)
            }

            if self.save_intermediate:
                self._save_intermediate('04_transformation_strategy.json',
                                        transformation_result.get('transformation_strategy', {}))
                self._save_intermediate('05_transformation_code.py',
                                        transformation_result['transformation_code'],
                                        is_text=True)

            # ================================================================
            # Stage 5: SAVE OUTPUT
            # ================================================================
            logger.info("\n" + "=" * 80)
            logger.info("STAGE 5: SAVING OUTPUT")
            logger.info("=" * 80)

            if output_file is None:
                output_file = self.output_dir / 'normalized_output.csv'

            output_path = self._save_output(transformation_result['normalized_df'], output_file)
            pipeline_log['output_file'] = str(output_path)

            # Calculate execution time
            elapsed_time = time.time() - start_time
            pipeline_log['end_time'] = time.strftime('%Y-%m-%d %H:%M:%S')
            pipeline_log['elapsed_seconds'] = round(elapsed_time, 2)
            pipeline_log['status'] = 'success'

            # Print summary
            self._print_summary(pipeline_log, transformation_result, encoded_data)

            # Save pipeline log
            if self.save_intermediate:
                self._save_intermediate('pipeline_log.json', pipeline_log)

            return {
                'normalized_df': transformation_result['normalized_df'],
                'output_path': output_path,
                'pipeline_log': pipeline_log,
                'intermediate_results': {
                    'encoded_data': encoded_data,
                    'structure_analysis': structure_analysis,
                    'schema': schema,
                    'transformation': transformation_result
                }
            }

        except Exception as e:
            logger.error(f"\n{'=' * 80}")
            logger.error(f"PIPELINE FAILED: {e}")
            logger.error(f"{'=' * 80}\n")

            pipeline_log['status'] = 'failed'
            pipeline_log['error'] = str(e)
            pipeline_log['elapsed_seconds'] = round(time.time() - start_time, 2)

            if self.save_intermediate:
                self._save_intermediate('pipeline_log.json', pipeline_log)

            raise

    # ========== SMART TABLE DETECTION ==========
    
    def _detect_tables_smart(self, input_file: str) -> List[Dict[str, Any]]:
        """
        Smart table detection: try hyperlinks first, then fallback to empty row.
        """
        file_path = Path(input_file)
        
        logger.info(f"Detection mode: {self.detection_mode}")
        
        # Mode 1: Auto - try hyperlink first for .xlsx files
        if self.detection_mode == 'auto':
            if file_path.suffix.lower() == '.xlsx':
                try:
                    tables = self._detect_tables_by_hyperlink(str(file_path))
                    
                    if len(tables) > 1:
                        logger.info(f"✓ Hyperlink detection succeeded: {len(tables)} tables")
                        for table in tables:
                            table['detection_method'] = 'hyperlink'
                        return tables
                    else:
                        logger.info("Only one table via hyperlinks, trying empty row detection")
                except Exception as e:
                    logger.warning(f"Hyperlink detection failed: {e}, trying empty row detection")
            
            # Fallback to empty row detection
            return self._detect_tables_by_empty_rows(input_file)
        
        # Mode 2: Force hyperlink
        elif self.detection_mode == 'hyperlink':
            tables = self._detect_tables_by_hyperlink(str(file_path))
            for table in tables:
                table['detection_method'] = 'hyperlink'
            return tables
        
        # Mode 3: Force empty row
        elif self.detection_mode == 'empty_row':
            return self._detect_tables_by_empty_rows(input_file)
        
        else:
            raise ValueError(f"Unknown detection mode: {self.detection_mode}")
    
    # ========== HYPERLINK DETECTION (BUILT-IN) ==========
    
    def _detect_tables_by_hyperlink(self, file_path: str) -> List[Dict[str, Any]]:
        """
        Detect tables by following hyperlinks in Excel file.
        Built-in implementation - no separate module needed.
        """
        file_path = Path(file_path)
        
        if file_path.suffix.lower() != '.xlsx':
            logger.warning("Hyperlink detection only works with .xlsx files")
            df = pd.read_excel(file_path)
            return [{
                'table_id': 1,
                'table_name': 'Table_1',
                'source': 'main',
                'dataframe': df
            }]
        
        logger.info(f"Detecting tables via hyperlinks in: {file_path}")
        
        # Load workbook
        wb = openpyxl.load_workbook(file_path, data_only=False)
        
        # Find main sheet
        if self.main_sheet_name:
            if self.main_sheet_name not in wb.sheetnames:
                raise ValueError(f"Main sheet '{self.main_sheet_name}' not found")
            main_sheet = wb[self.main_sheet_name]
        else:
            main_sheet = wb.worksheets[0]
        
        logger.info(f"Main sheet: '{main_sheet.title}'")
        
        # Extract hyperlinks
        hyperlinks = self._extract_hyperlinks_from_sheet(main_sheet)
        
        if not hyperlinks:
            logger.info("No hyperlinks found, loading main sheet as single table")
            df = pd.read_excel(file_path, sheet_name=main_sheet.title)
            return [{
                'table_id': 1,
                'table_name': main_sheet.title,
                'source': 'main',
                'dataframe': df
            }]
        
        logger.info(f"Found {len(hyperlinks)} hyperlink(s)")
        
        # Load main sheet + linked sheets
        tables = []
        
        # Add main sheet
        main_df = pd.read_excel(file_path, sheet_name=main_sheet.title)
        tables.append({
            'table_id': 1,
            'table_name': main_sheet.title,
            'source': 'main',
            'sheet_name': main_sheet.title,
            'dataframe': main_df,
            'hyperlink_info': None
        })
        
        # Add linked sheets
        for idx, link_info in enumerate(hyperlinks, start=2):
            if link_info['type'] == 'internal':
                try:
                    linked_df = pd.read_excel(file_path, sheet_name=link_info['target_sheet'])
                    tables.append({
                        'table_id': idx,
                        'table_name': link_info['target_sheet'],
                        'source': 'hyperlink',
                        'sheet_name': link_info['target_sheet'],
                        'dataframe': linked_df,
                        'hyperlink_info': link_info
                    })
                    logger.info(f"  Loaded linked sheet: '{link_info['target_sheet']}' ({linked_df.shape})")
                except Exception as e:
                    logger.error(f"  Failed to load sheet '{link_info['target_sheet']}': {e}")
            
            elif link_info['type'] == 'external' and self.follow_external:
                try:
                    external_file = Path(file_path).parent / link_info['target_file']
                    if external_file.exists():
                        external_df = pd.read_excel(external_file)
                        tables.append({
                            'table_id': idx,
                            'table_name': external_file.stem,
                            'source': 'external',
                            'file_path': str(external_file),
                            'dataframe': external_df,
                            'hyperlink_info': link_info
                        })
                        logger.info(f"  Loaded external file: '{external_file.name}' ({external_df.shape})")
                except Exception as e:
                    logger.error(f"  Failed to load external file '{link_info['target_file']}': {e}")
        
        logger.info(f"Total tables detected: {len(tables)}")
        
        return tables
    
    def _extract_hyperlinks_from_sheet(self, worksheet) -> List[Dict[str, Any]]:
        """Extract all hyperlinks from a worksheet that point to OTHER sheets."""
        hyperlinks = []
        seen_targets = set()
        
        for row in worksheet.iter_rows():
            for cell in row:
                if cell.hyperlink:
                    link_info = self._parse_hyperlink(cell)
                    
                    if link_info and link_info['type'] == 'internal':
                        target_sheet = link_info['target_sheet']
                        
                        # Skip if pointing to the same sheet (internal bookmark)
                        if target_sheet == worksheet.title:
                            continue
                        
                        # Avoid duplicate targets
                        if target_sheet not in seen_targets:
                            hyperlinks.append(link_info)
                            seen_targets.add(target_sheet)
                            
                            logger.debug(
                                f"  Found link at {cell.coordinate}: "
                                f"'{cell.value}' → '{target_sheet}'"
                            )
        
        return hyperlinks
    
    def _parse_hyperlink(self, cell) -> Optional[Dict[str, Any]]:
        """Parse a cell's hyperlink and return structured info."""
        if not cell.hyperlink:
            return None
        
        # Try both target and location attributes
        # Some Excel files use target, others use location
        target = cell.hyperlink.target
        location = getattr(cell.hyperlink, 'location', None)
        
        # Use whichever is available
        link_ref = target or location
        
        if not link_ref:
            return None
        
        # Internal sheet link: #SheetName!A1 or SheetName!A1
        if '!' in link_ref:
            # Remove leading # if present
            if link_ref.startswith('#'):
                link_ref = link_ref[1:]
            
            parts = link_ref.split('!')
            
            if len(parts) >= 1:
                sheet_name = parts[0].strip("'\"")
                cell_ref = parts[1] if len(parts) > 1 else None
                
                return {
                    'type': 'internal',
                    'from_cell': cell.coordinate,
                    'cell_text': str(cell.value) if cell.value else '',
                    'target_sheet': sheet_name,
                    'target_cell': cell_ref,
                    'raw_target': link_ref
                }
        
        # External file link (only check target, not location)
        elif target and target.endswith(('.xlsx', '.xls', '.csv')):
            return {
                'type': 'external',
                'from_cell': cell.coordinate,
                'cell_text': str(cell.value) if cell.value else '',
                'target_file': target,
                'raw_target': target
            }
        
        # HTTP/web links - ignore (only check target)
        elif target and target.startswith(('http://', 'https://')):
            return None
        
        return None
    
    # ========== EMPTY ROW DETECTION (BUILT-IN) ==========
    
    def _detect_tables_by_empty_rows(self, input_file: str) -> List[Dict[str, Any]]:
        """
        Detect tables by empty row separation (original method).
        Built-in implementation.
        """
        df = self.encoder.load_file(input_file)
        
        logger.info(f"Detecting tables by empty rows in DataFrame of shape {df.shape}")
        
        # Detect table regions
        detected_tables = self._detect_table_regions(df)
        
        # Split DataFrame
        split_tables = self._split_dataframe(df, detected_tables)
        
        # Add detection method
        for table in split_tables:
            table['detection_method'] = 'empty_row'
            table['source'] = 'empty_row_detection'
        
        # Visualize
        viz = self._visualize_detection(df, detected_tables)
        logger.info(f"\n{viz}")
        
        if self.save_intermediate:
            self._save_intermediate('detected_tables.txt', viz, is_text=True)
        
        return split_tables
    
    def _detect_table_regions(self, df: pd.DataFrame) -> List[TableRegion]:
        """Detect multiple table regions within a DataFrame using empty row separators."""
        tables = []
        
        # Calculate emptiness for each row
        row_emptiness = df.isna().mean(axis=1)
        
        # Find separator rows
        separator_rows = row_emptiness >= self.empty_threshold
        
        # Find contiguous data regions
        current_start = None
        
        for idx in range(len(df)):
            is_separator = separator_rows.iloc[idx]
            
            if not is_separator and current_start is None:
                current_start = idx
            elif is_separator and current_start is not None:
                if idx - current_start >= self.min_table_rows:
                    tables.append(self._create_table_region(
                        df, current_start, idx - 1, 0, len(df.columns) - 1,
                        f"Table_{len(tables) + 1}"
                    ))
                current_start = None
        
        # Handle last table
        if current_start is not None:
            if len(df) - current_start >= self.min_table_rows:
                tables.append(self._create_table_region(
                    df, current_start, len(df) - 1, 0, len(df.columns) - 1,
                    f"Table_{len(tables) + 1}"
                ))
        
        # If no tables detected, return entire DataFrame as one table
        if not tables:
            tables.append(self._create_table_region(
                df, 0, len(df) - 1, 0, len(df.columns) - 1, "Table_1"
            ))
        
        logger.info(f"Detected {len(tables)} table(s)")
        for i, table in enumerate(tables):
            logger.info(f"  {table.name}: rows [{table.start_row}:{table.end_row}], "
                       f"shape {table.end_row - table.start_row + 1} x "
                       f"{table.end_col - table.start_col + 1}, "
                       f"confidence {table.confidence:.2%}")
        
        return tables
    
    def _create_table_region(self, df: pd.DataFrame, 
                            start_row: int, end_row: int,
                            start_col: int, end_col: int,
                            name: str) -> TableRegion:
        """Create a TableRegion with confidence score."""
        region_df = df.iloc[start_row:end_row + 1, start_col:end_col + 1]
        data_density = 1 - region_df.isna().mean().mean()
        
        return TableRegion(
            start_row=start_row,
            end_row=end_row,
            start_col=start_col,
            end_col=end_col,
            name=name,
            confidence=float(data_density)
        )
    
    def _split_dataframe(self, df: pd.DataFrame, tables: List[TableRegion]) -> List[Dict[str, Any]]:
        """Split a DataFrame into multiple DataFrames based on detected table regions."""
        split_dfs = []
        
        for i, table in enumerate(tables):
            table_df = df.iloc[
                table.start_row:table.end_row + 1,
                table.start_col:table.end_col + 1
            ].copy()
            
            table_df = table_df.reset_index(drop=True)
            
            split_dfs.append({
                'table_id': i + 1,
                'table_name': table.name,
                'dataframe': table_df,
                'region': table.to_dict(),
                'original_position': {
                    'row_range': (table.start_row, table.end_row),
                    'col_range': (table.start_col, table.end_col)
                }
            })
            
            logger.info(f"Extracted {table.name}: shape {table_df.shape}")
        
        return split_dfs
    
    def _visualize_detection(self, df: pd.DataFrame, tables: List[TableRegion]) -> str:
        """Generate a text visualization of detected tables."""
        lines = [
            f"DataFrame shape: {df.shape}",
            f"Detected {len(tables)} table(s):",
            ""
        ]
        
        for i, table in enumerate(tables):
            lines.append(f"{table.name}:")
            lines.append(f"  Position: rows [{table.start_row}:{table.end_row}], "
                        f"cols [{table.start_col}:{table.end_col}]")
            lines.append(f"  Shape: {table.end_row - table.start_row + 1} x "
                        f"{table.end_col - table.start_col + 1}")
            lines.append(f"  Confidence: {table.confidence:.2%}")
            lines.append("")
        
        return "\n".join(lines)
    
    # ========== MULTI-TABLE PROCESSING ==========

    def _process_multiple_tables(self, 
                                 split_tables: List[Dict[str, Any]],
                                 pipeline_log: Dict[str, Any],
                                 start_time: float,
                                 output_file: Optional[str]) -> Dict[str, Any]:
        """Process multiple detected tables either separately or combined."""
        logger.info(f"\nProcessing mode: {self.split_processing_mode}")
        logger.info(f"Detection method: {split_tables[0].get('detection_method', 'unknown')}")
        
        all_results = []
        all_encoded_data = []
        for table_info in split_tables:
            table_id = table_info['table_id']
            table_name = table_info['table_name']
            table_df = table_info['dataframe']
            
            logger.info(f"\n{'='*80}")
            logger.info(f"PROCESSING {table_name} (ID: {table_id})")
            logger.info(f"Shape: {table_df.shape}")
            logger.info(f"Source: {table_info.get('source', 'unknown')}")
            logger.info(f"{'='*80}")
            
            try:
                # For each table, create a fresh temporary file
                # This ensures each table has its own working file
                temp_file = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
                temp_file_path = temp_file.name
                temp_file.close()
                
                # Write this table's DataFrame to the temp file
                table_df.to_excel(temp_file_path, index=False, engine='openpyxl')
                
                # Set encoder to use this temp file
                self.encoder._working_file = temp_file_path
                self.encoder._temp_xlsx = temp_file_path
                
                # Now encode
                table_encoded = self.encoder.encode(table_df)
                table_encoded['table_id'] = table_id
                table_encoded['table_name'] = table_name
                table_encoded['source'] = table_info.get('source')
                encoded_for_summary = {
                        k: v for k, v in table_encoded.items() if k != 'dataframe'
                    }
                if 'encoded_text' in encoded_for_summary and len(encoded_for_summary['encoded_text']) > 500:
                    encoded_for_summary['encoded_text'] = encoded_for_summary['encoded_text'][:500] + '...'
                all_encoded_data.append(encoded_for_summary)
                # Run the pipeline for this table
                structure_analysis = self._run_analysis(table_encoded)
                schema = self._run_schema_estimation(table_encoded, structure_analysis)
                transformation = self._run_transformation(table_encoded, structure_analysis, schema)
                
                result = {
                    'table_id': table_id,
                    'table_name': table_name,
                    'source': table_info.get('source'),
                    'normalized_df': transformation['normalized_df'],
                    'region': table_info.get('region'),
                    'structure_analysis': structure_analysis,
                    'schema': schema,
                    'validation': transformation['validation_result'],
                    'hyperlink_info': table_info.get('hyperlink_info')
                }
                
                all_results.append(result)
                
                # Save individual table output
                if self.save_intermediate:
                    table_output = self.output_dir / f'{table_name}_normalized.csv'
                    self._save_output(transformation['normalized_df'], table_output)
                    self._save_intermediate(
                        f'{table_name}_transformation.py',
                        transformation['transformation_code'],
                        is_text=True
                    )
                encoded_data_to_save = {
                    k: v for k, v in table_encoded.items() if k != 'dataframe'
                }
                if 'encoded_text' in encoded_data_to_save and len(encoded_data_to_save['encoded_text']) > 1000:
                    encoded_data_to_save['encoded_text'] = encoded_data_to_save['encoded_text'][:1000] + '...'
                
                self._save_intermediate(f'{table_name}_encoded_data.json', encoded_data_to_save)
                logger.info(f"Saved encoded data for {table_name}")
                logger.info(f"{table_name} processed successfully")
                
                # Clean up the temp file for this table
                try:
                    if Path(temp_file_path).exists():
                        Path(temp_file_path).unlink()
                except Exception as cleanup_error:
                    logger.warning(f"Could not delete temp file {temp_file_path}: {cleanup_error}")
                
            except Exception as e:
                logger.error(f"Failed to process {table_name}: {e}")
                
                # Still try to clean up temp file even if processing failed
                try:
                    if 'temp_file_path' in locals() and Path(temp_file_path).exists():
                        Path(temp_file_path).unlink()
                except Exception:
                    pass
                
                result = {
                    'table_id': table_id,
                    'table_name': table_name,
                    'source': table_info.get('source'),
                    'error': str(e),
                    'region': table_info.get('region')
                }
                all_results.append(result)
        
        # Combine results if needed
        if self.split_processing_mode == 'combined':
            combined_df = self._combine_table_results(all_results)
            output_path = self._save_output(
                combined_df, 
                output_file or self.output_dir / 'combined_normalized.csv'
            )
        else:
            output_path = self.output_dir / 'tables_summary.json'
            summary = {
                'num_tables': len(split_tables),
                'detection_method': split_tables[0].get('detection_method', 'unknown'),
                'tables': [
                    {
                        'id': r['table_id'],
                        'name': r['table_name'],
                        'source': r.get('source'),
                        'shape': r['normalized_df'].shape if 'normalized_df' in r else None,
                        'status': 'success' if 'normalized_df' in r else 'failed',
                        'hyperlink_info': r.get('hyperlink_info')
                    }
                    for r in all_results
                ]
            }
            if self.save_intermediate:
                self._save_intermediate(output_path.name, summary)
                if all_encoded_data:
                    combined = {
                        'num_tables': len(all_encoded_data),
                        'detection_method': split_tables[0].get('detection_method', 'unknown'),
                        'tables': all_encoded_data
                    }
                    self._save_intermediate('all_tables_encoded_data.json', combined)
                    logger.info(f"Saved combined encoded data for {len(all_encoded_data)} tables")
        
        # Update pipeline log
        elapsed_time = time.time() - start_time
        pipeline_log['end_time'] = time.strftime('%Y-%m-%d %H:%M:%S')
        pipeline_log['elapsed_seconds'] = round(elapsed_time, 2)
        pipeline_log['processing_mode'] = self.split_processing_mode
        pipeline_log['detection_method'] = split_tables[0].get('detection_method', 'unknown')
        pipeline_log['tables_processed'] = len(all_results)
        
        self._save_intermediate('pipeline_log.json', pipeline_log)
        
        logger.info(f"\n{'='*80}")
        logger.info(f"ALL TABLES PROCESSED in {elapsed_time:.2f} seconds")
        logger.info(f"Detection method: {split_tables[0].get('detection_method', 'unknown')}")
        logger.info(f"Output saved to: {output_path}")
        logger.info(f"{'='*80}\n")
        
        return {
            'all_tables': all_results,
            'output_path': output_path,
            'pipeline_log': pipeline_log,
            'num_tables': len(split_tables),
            'detection_method': split_tables[0].get('detection_method', 'unknown')
        }

    def _combine_table_results(self, results: List[Dict[str, Any]]) -> pd.DataFrame:
        """Combine multiple normalized tables into a single DataFrame."""
        valid_results = [r for r in results if 'normalized_df' in r]
        
        if not valid_results:
            raise ValueError("No valid results to combine")
        
        # Add table identifier to each DataFrame
        for result in valid_results:
            result['normalized_df']['source_table'] = result['table_name']
            result['normalized_df']['source_table_id'] = result['table_id']
        
        # Concatenate
        combined = pd.concat(
            [r['normalized_df'] for r in valid_results],
            ignore_index=True
        )
        
        logger.info(f"Combined {len(valid_results)} tables into shape {combined.shape}")
        
        return combined

    # ========== PIPELINE COMPONENTS ==========

    def _run_encoding(self, input_file: str) -> Dict[str, Any]:
        """Run the encoding stage."""
        df = self.encoder.load_file(input_file)
        logger.info(f"Loaded data with shape: {df.shape}")

        encoded_data = self.encoder.encode(df)
        logger.info(f"Encoding complete. Compression: {encoded_data['metadata'].get('compression_ratio', 0):.2f}x")

        return encoded_data

    def _run_structure_analysis(self, encoded_data: Dict[str, Any]) -> Dict[str, Any]:
        """Run the structure analysis stage with LLM semantic reasoning."""
        analysis = self.analyzer.analyze(encoded_data)

        # Log key findings
        sem = analysis.get('semantic_understanding', {})
        logger.info(f"Data description: {sem.get('data_description', 'Unknown')[:100]}")

        row_patterns = analysis.get('row_patterns', {})
        if row_patterns.get('has_bilingual_rows'):
            bilingual = row_patterns.get('bilingual_details', {})
            logger.info(f"Bilingual pattern: {bilingual.get('data_relationship', 'unknown')}")

        if row_patterns.get('has_section_markers'):
            markers = row_patterns.get('section_marker_details', {})
            logger.info(f"Section markers: {markers.get('marker_values', [])}")

        implicit_agg = analysis.get('implicit_aggregation', {})
        if implicit_agg.get('has_implicit_aggregation'):
            logger.info(f"Implicit aggregation detected: {len(implicit_agg.get('aggregation_hierarchies', []))} hierarchies")
            logger.info(f"  Summary rows to exclude: {len(implicit_agg.get('summary_rows', []))}")

        logger.info(f"Transformation complexity: {analysis.get('transformation_complexity', 'unknown')}")

        return analysis

    def _run_schema_estimation(self,
                               encoded_data: Dict[str, Any],
                               structure_analysis: Dict[str, Any]) -> Dict[str, Any]:
        """Run the schema estimation stage based on tidy data principles."""
        schema = self.estimator.estimate_schema(encoded_data, structure_analysis)

        # Log target schema
        obs_unit = schema.get('observation_unit', {})
        logger.info(f"Observation unit: {obs_unit.get('description', 'Unknown')}")

        target_cols = schema.get('target_columns', [])
        logger.info(f"Target columns ({len(target_cols)}):")
        for col in target_cols:
            dim_marker = "[DIM]" if col.get('is_dimension') else "[VAL]"
            logger.info(f"  {dim_marker} {col.get('name')} ({col.get('data_type')})")

        expected = schema.get('expected_output', {})
        logger.info(f"Expected output: {expected.get('row_count_estimate', '?')} rows")

        return schema

    def _run_transformation(self,
                            encoded_data: Dict[str, Any],
                            structure_analysis: Dict[str, Any],
                            schema: Dict[str, Any]) -> Dict[str, Any]:
        """Run the two-stage transformation generation and execution."""
        result = self.generator.generate_and_execute(
            encoded_data, structure_analysis, schema
        )

        # Log results
        logger.info(f"Transformation completed in {result.get('attempts', 1)} attempt(s)")
        logger.info(f"Output shape: {result['normalized_df'].shape}")

        validation = result['validation_result']
        if validation['is_valid']:
            logger.info("✓ Validation PASSED")
        else:
            logger.warning(f"✗ Validation issues: {validation.get('errors', [])}")

        if validation.get('warnings'):
            for warning in validation['warnings']:
                logger.warning(f"  Warning: {warning}")

        return result

    def _save_output(self, df: pd.DataFrame, output_file: str) -> Path:
        """Save the normalized DataFrame to file."""
        output_path = Path(output_file)
        output_path.parent.mkdir(parents=True, exist_ok=True)

        output_format = self.output_config.get('format', 'csv')
        encoding = self.output_config.get('encoding', 'utf-8')
        include_index = self.output_config.get('index', False)

        if output_format == 'csv' or output_path.suffix == '.csv':
            df.to_csv(output_path, index=include_index, encoding=encoding)
        elif output_format == 'xlsx' or output_path.suffix == '.xlsx':
            df.to_excel(output_path, index=include_index, engine='openpyxl')
        else:
            df.to_csv(output_path, index=include_index, encoding=encoding)

        logger.info(f"Saved output to: {output_path}")
        return output_path

    def _save_intermediate(self, filename: str, data: Any, is_text: bool = False):
        """Save intermediate results to the output directory."""
        try:
            filepath = self.output_dir / filename

            if is_text:
                with open(filepath, 'w', encoding='utf-8') as f:
                    f.write(data)
            else:
                with open(filepath, 'w', encoding='utf-8') as f:
                    json.dump(data, f, indent=2, ensure_ascii=False, default=str)

            logger.debug(f"Saved intermediate result: {filename}")
        except Exception as e:
            logger.warning(f"Failed to save intermediate result {filename}: {e}")

    def _print_summary(self,
                       pipeline_log: Dict[str, Any],
                       transformation_result: Dict[str, Any],
                       encoded_data: Dict[str, Any]):
        """Print final pipeline summary."""
        logger.info("\n" + "=" * 80)
        logger.info("NORMALIZATION SUMMARY")
        logger.info("=" * 80)

        logger.info(f"Status: {pipeline_log.get('status', 'unknown').upper()}")
        logger.info(f"Input shape: {encoded_data['original_shape']}")
        logger.info(f"Output shape: {transformation_result['normalized_df'].shape}")
        logger.info(f"Compression: {encoded_data['metadata'].get('compression_ratio', 0):.2f}x")
        logger.info(f"Output file: {pipeline_log.get('output_file', 'N/A')}")
        logger.info(f"Execution time: {pipeline_log.get('elapsed_seconds', 0)} seconds")
        logger.info(f"Transformation attempts: {transformation_result.get('attempts', 1)}")

        validation = transformation_result['validation_result']
        logger.info(f"Validation: {'PASSED' if validation['is_valid'] else 'FAILED'}")

        if validation.get('warnings'):
            logger.info(f"Warnings: {len(validation['warnings'])}")
            for w in validation['warnings'][:3]:
                logger.info(f"  - {w}")

        logger.info("=" * 80 + "\n")
