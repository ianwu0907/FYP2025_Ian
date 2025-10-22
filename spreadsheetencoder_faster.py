#!/usr/bin/env python3
"""
spreadsheetllm_runner.py
Single-file runner: helpers + adapted SpreadsheetLLM encoder.
Usage:
    python spreadsheetllm_runner.py /path/to/input.xlsx -o /path/to/output.json
    python spreadsheetllm_runner.py /path/to/input.csv  -o /path/to/output.json
"""
import os
import json
import argparse
import logging
import re
import datetime
from collections import defaultdict
from pathlib import Path

import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string

# Optional: pandas for CSV->XLSX conversion
try:
    import pandas as pd
except Exception:
    pd = None

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger("spreadsheetllm_runner")

# -----------------------
# User-provided helpers
# -----------------------
EMAIL_REGEX = re.compile(r"^[a-zA-Z0-9_.+-]+@[a-zA-Z0-9-]+\.[a-zA-Z0-9-.]+$")

def infer_cell_data_type(cell: openpyxl.cell.cell.Cell) -> str:
    if cell.value is None:
        return "empty"
    if isinstance(cell.value, str) and EMAIL_REGEX.match(cell.value):
        return "email"
    data_type = getattr(cell, "data_type", None)
    if data_type == 's':
        return "text"
    elif data_type == 'n':
        return "numeric"
    elif data_type == 'b':
        return "boolean"
    elif data_type == 'd':
        return "datetime"
    elif data_type == 'e':
        return "error"
    elif data_type == 'f':
        if cell.value is not None:
            if isinstance(cell.value, str):
                return "text"
            elif isinstance(cell.value, (int, float)):
                return "numeric"
            elif isinstance(cell.value, bool):
                return "boolean"
            else:
                return "formula_cached_value"
        return "formula"
    elif data_type == 'g':
        if cell.value is not None:
            return "text"
        else:
            return "empty"
    else:
        return "unknown"

def categorize_number_format(number_format_string: str, cell: openpyxl.cell.cell.Cell) -> str:
    cell_data_type = infer_cell_data_type(cell)
    if cell_data_type not in ["numeric", "datetime"]:
        return "not_applicable"
    if number_format_string is None or str(number_format_string).lower() == "general":
        if cell_data_type == "datetime":
            return "datetime_general"
        return "general"
    if number_format_string == "@" or str(number_format_string).lower() == "text":
        return "text_format"
    if any(c in str(number_format_string) for c in ['$', '€', '£', '¥']):
        return "currency"
    if '%' in str(number_format_string):
        return "percentage"
    nf_lower = str(number_format_string).lower()
    if 'e+' in nf_lower or 'e-' in nf_lower:
        return "scientific"
    if '#' in nf_lower and '/' in nf_lower and '?' in nf_lower:
        return "fraction"
    date_keywords = ['yyyy', 'yy', 'mmmm', 'mmm', 'mm', 'dddd', 'ddd', 'dd', 'd']
    time_keywords = ['hh', 'h', 'ss', 's', 'am/pm', 'a/p']
    is_date = any(k in nf_lower for k in date_keywords)
    is_time = any(k in nf_lower for k in time_keywords)
    if ':' in str(number_format_string):
        tmp = str(number_format_string).replace('0','').replace('#','').replace(',','').replace('.','')
        if ':' in tmp:
            is_time = True
    if is_date and is_time:
        return "datetime_custom"
    if is_date:
        return "date_custom"
    if is_time:
        return "time_custom"
    if cell_data_type == "numeric":
        if number_format_string in ["0", "#,##0"]:
            return "integer"
        if number_format_string in ["0.00", "#,##0.00", "0.0", "#,##0.0"]:
            return "float"
        return "other_numeric"
    if cell_data_type == "datetime":
        return "other_date"
    return "unknown_format_category"

def get_number_format_string(cell: openpyxl.cell.cell.Cell) -> str:
    try:
        nfs = cell.number_format
        if nfs is None or nfs == "":
            return "General"
        return str(nfs)
    except Exception:
        return "General"

def detect_semantic_type(cell: openpyxl.cell.cell.Cell) -> str:
    data_type = infer_cell_data_type(cell)
    if data_type == "email":
        return "email"
    nfs = get_number_format_string(cell)
    category = categorize_number_format(nfs, cell)
    nfs_lower = nfs.lower()
    if category == "percentage":
        return "percentage"
    if category == "currency":
        return "currency"
    if category in ["date_custom", "datetime_custom", "datetime_general", "other_date"]:
        if ("yyyy" in nfs_lower or "yy" in nfs_lower) and not any(x in nfs_lower for x in ["m", "d"]):
            return "year"
        return "date"
    if category == "time_custom":
        return "time"
    if category == "scientific":
        return "scientific_notation"
    if data_type == "numeric":
        if isinstance(cell.value, int) or category == "integer":
            return "integer"
        if isinstance(cell.value, float) or category == "float":
            return "float"
        return "numeric"
    return data_type

# -----------------------
# Utility / encoding code
# -----------------------
def split_cell_ref(cell_ref):
    col_str = ''.join(filter(str.isalpha, cell_ref))
    row_str = ''.join(filter(str.isdigit, cell_ref))
    return col_str, int(row_str)

def _merge_refs(refs):
    coords = []
    for ref in sorted(set(refs)):
        try:
            col_letter, row = split_cell_ref(ref)
            col = column_index_from_string(col_letter)
            coords.append((row, col))
        except Exception:
            continue
    cell_set = set(coords)
    processed = set()
    ranges = []
    for row, col in sorted(coords):
        if (row, col) in processed:
            continue
        width = 1
        while (row, col + width) in cell_set and (row, col + width) not in processed:
            width += 1
        height = 1
        expanding = True
        while expanding:
            next_row = row + height
            for w in range(width):
                if (next_row, col + w) not in cell_set or (next_row, col + w) in processed:
                    expanding = False
                    break
            if expanding:
                height += 1
        end_col = col + width - 1
        end_row = row + height - 1
        start_ref = f"{get_column_letter(col)}{row}"
        end_ref = f"{get_column_letter(end_col)}{end_row}"
        if width == 1 and height == 1:
            ranges.append(start_ref)
        else:
            ranges.append(f"{start_ref}:{end_ref}")
        for r in range(row, row + height):
            for c in range(col, col + width):
                processed.add((r, c))
    return ranges

def aggregate_formats(sheet, format_map):
    aggregated_formats = defaultdict(list)
    processed_cells = set()
    type_nfs_map = defaultdict(list)
    for _, cells in format_map.items():
        for cell_ref in cells:
            try:
                cell = sheet[cell_ref]
            except Exception:
                continue
            nfs = get_number_format_string(cell)
            sem_type = detect_semantic_type(cell)
            key = json.dumps({"type": sem_type, "nfs": nfs}, sort_keys=True)
            type_nfs_map[key].append(cell_ref)
    for key, cells in type_nfs_map.items():
        cells_set = set(cells)
        for start_cell in cells:
            if start_cell in processed_cells:
                continue
            try:
                start_col_letter, start_row = split_cell_ref(start_cell)
                start_col = column_index_from_string(start_col_letter)
            except Exception:
                continue
            best_width = 1
            best_height = 1
            best_area = 1
            best_end_cell = start_cell
            max_width = min(20, sheet.max_column - start_col + 1)
            max_height = min(20, sheet.max_row - start_row + 1)
            for width in range(1, max_width + 1):
                for height in range(1, max_height + 1):
                    valid_rectangle = True
                    for r in range(start_row, start_row + height):
                        for c in range(start_col, start_col + width):
                            cell_ref = f"{get_column_letter(c)}{r}"
                            if cell_ref not in cells_set or cell_ref in processed_cells:
                                valid_rectangle = False
                                break
                        if not valid_rectangle:
                            break
                    if valid_rectangle:
                        area = width * height
                        if area > best_area:
                            best_width = width
                            best_height = height
                            best_area = area
                            best_end_cell = f"{get_column_letter(start_col + width - 1)}{start_row + height - 1}"
            region = start_cell if best_width == 1 and best_height == 1 else f"{start_cell}:{best_end_cell}"
            aggregated_formats[key].append(region)
            for r in range(start_row, start_row + best_height):
                for c in range(start_col, start_col + best_width):
                    processed_cells.add(f"{get_column_letter(c)}{r}")
    return dict(aggregated_formats)

# fast anchors heuristic (safe & fast)
def fast_find_structural_anchors(sheet, k=2):
    header_rows = []
    maxr = min(21, sheet.max_row)
    maxc = min(21, sheet.max_column)
    for r in range(1, maxr + 1):
        populated = 0
        strings = 0
        for c in range(1, maxc + 1):
            cell = sheet.cell(row=r, column=c)
            if cell.value is not None and str(cell.value).strip() != "":
                populated += 1
                if isinstance(cell.value, str):
                    strings += 1
        if populated >= 1 and strings / (populated or 1) > 0.5:
            header_rows.append(r)
    if not header_rows:
        header_rows = [1]
    cols = []
    for c in range(1, sheet.max_column + 1):
        found = False
        for r in range(1, min(51, sheet.max_row + 1)):
            if sheet.cell(row=r, column=c).value is not None and str(sheet.cell(row=r, column=c).value).strip() != "":
                found = True
                break
        if found:
            cols.append(c)
    if not cols:
        cols = list(range(1, min(6, sheet.max_column + 1)))
    def expand_idxs(idxs, k, max_idx):
        s = set()
        for idx in idxs:
            for i in range(max(1, idx - k), min(max_idx + 1, idx + k + 1)):
                s.add(i)
        return sorted(s)
    row_anchors = expand_idxs(header_rows, k, sheet.max_row)
    col_anchors = expand_idxs(cols[:max(1, min(10, len(cols)))], k, sheet.max_column)
    return row_anchors, col_anchors

def compress_homogeneous_regions(sheet, rows, cols):
    def row_homogeneous(r):
        vals = []
        fmts = []
        for c in cols:
            cell = sheet.cell(row=r, column=c)
            vals.append(cell.value)
            fmts.append(get_number_format_string(cell))
        return len(set(vals)) <= 1 and len(set(fmts)) <= 1
    def col_homogeneous(c):
        vals = []
        fmts = []
        for r in rows:
            cell = sheet.cell(row=r, column=c)
            vals.append(cell.value)
            fmts.append(get_number_format_string(cell))
        return len(set(vals)) <= 1 and len(set(fmts)) <= 1
    filtered_rows = [r for r in rows if not row_homogeneous(r)]
    filtered_cols = [c for c in cols if not col_homogeneous(c)]
    return filtered_rows, filtered_cols

def create_inverted_index(sheet, kept_rows, kept_cols):
    inverted_index = defaultdict(list)
    format_map = defaultdict(list)
    merged_ranges = sheet.merged_cells.ranges
    for row in kept_rows:
        for col in kept_cols:
            cell = sheet.cell(row=row, column=col)
            cell_ref = f"{get_column_letter(col)}{row}"
            merged_value = None
            merged_range = None
            for m_range in merged_ranges:
                if cell_ref in m_range:
                    try:
                        merged_value = sheet[m_range.start_cell.coordinate].value
                        merged_range = m_range
                        break
                    except Exception:
                        pass
            try:
                if merged_value is not None:
                    cell_value = str(merged_value) if merged_value is not None else ""
                    inverted_index[cell_value].append(cell_ref)
                elif cell.value is not None:
                    if isinstance(cell.value, (int, float)):
                        cell_value = f"{cell.value}"
                    else:
                        cell_value = str(cell.value)
                    inverted_index[cell_value].append(cell_ref)
            except Exception:
                inverted_index["ERROR_VALUE"].append(cell_ref)
            try:
                format_info = {}
                format_info["font"] = {"bold": getattr(cell.font, 'bold', None)}
                alignment = getattr(cell, 'alignment', None)
                format_info["alignment"] = {"horizontal": getattr(alignment, 'horizontal', None)}
                original_number_format = get_number_format_string(cell)
                inferred_type = infer_cell_data_type(cell)
                category = categorize_number_format(original_number_format, cell)
                format_info["original_number_format"] = original_number_format
                format_info["inferred_data_type"] = inferred_type
                format_info["number_format_category"] = category
                if merged_range is not None:
                    format_info["merged"] = True
                    format_info["merged_range"] = str(merged_range)
                else:
                    format_info["merged"] = False
                format_key = json.dumps(format_info, sort_keys=True)
                format_map[format_key].append(cell_ref)
            except Exception:
                pass
    return dict(inverted_index), dict(format_map)

def create_inverted_index_translation(inverted_index):
    merged_index = {}
    for value, refs in inverted_index.items():
        if value is None or str(value).strip() == "":
            continue
        merged_index[value] = _merge_refs(refs)
    return merged_index

# top-level encoding function
def spreadsheet_llm_encode_with_helpers(excel_path, output_path=None, k=2):
    wb = openpyxl.load_workbook(excel_path, data_only=False)
    sheets_encoding = {}
    compression_metrics = {"sheets": {}}
    overall_orig = overall_anchor = overall_index = overall_format = overall_final = 0
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        if sheet.max_row <= 1 and sheet.max_column <= 1:
            continue
        original_cells = {}
        for r in range(1, sheet.max_row + 1):
            for c in range(1, sheet.max_column + 1):
                v = sheet.cell(row=r, column=c).value
                if v is not None:
                    original_cells[f"{get_column_letter(c)}{r}"] = str(v)
        original_tokens = len(json.dumps(original_cells, ensure_ascii=False))
        row_anchors, col_anchors = fast_find_structural_anchors(sheet, k)
        kept_rows, kept_cols = extract_cells_near_anchors(sheet, row_anchors, col_anchors, 0)
        if not kept_rows or not kept_cols:
            kept_rows = list(range(1, min(21, sheet.max_row+1)))
            kept_cols = list(range(1, min(21, sheet.max_column+1)))
        kept_rows, kept_cols = compress_homogeneous_regions(sheet, kept_rows, kept_cols)
        anchor_cells = {}
        for r in kept_rows:
            for c in kept_cols:
                v = sheet.cell(row=r, column=c).value
                if v is not None:
                    anchor_cells[f"{get_column_letter(c)}{r}"] = str(v)
        anchor_tokens = len(json.dumps(anchor_cells, ensure_ascii=False))
        inverted_index, format_map = create_inverted_index(sheet, kept_rows, kept_cols)
        index_tokens = len(json.dumps(inverted_index, ensure_ascii=False))
        merged_index = create_inverted_index_translation(inverted_index)
        merged_tokens = len(json.dumps(merged_index, ensure_ascii=False))
        aggregated_formats = aggregate_formats(sheet, format_map)
        format_tokens = len(json.dumps(aggregated_formats, ensure_ascii=False))
        numeric_ranges = cluster_numeric_ranges(sheet, format_map)
        numeric_tokens = len(json.dumps(numeric_ranges, ensure_ascii=False))
        sheet_encoding = {"structural_anchors": {"rows": row_anchors, "columns": [get_column_letter(c) for c in col_anchors]}, "cells": merged_index, "formats": aggregated_formats, "numeric_ranges": numeric_ranges}
        final_tokens = len(json.dumps(sheet_encoding, ensure_ascii=False))
        compression_metrics["sheets"][sheet_name] = {"original_tokens": original_tokens, "after_anchor_tokens": anchor_tokens, "after_inverted_index_tokens": index_tokens, "after_merged_index_tokens": merged_tokens, "after_format_tokens": format_tokens, "numeric_tokens": numeric_tokens, "final_tokens": final_tokens, "anchor_ratio": (original_tokens / anchor_tokens) if anchor_tokens else 0, "index_ratio": (original_tokens / index_tokens) if index_tokens else 0, "format_ratio": (original_tokens / format_tokens) if format_tokens else 0, "overall_ratio": (original_tokens / final_tokens) if final_tokens else 0}
        sheets_encoding[sheet_name] = sheet_encoding
        overall_orig += original_tokens
        overall_anchor += anchor_tokens
        overall_index += index_tokens
        overall_format += format_tokens
        overall_final += final_tokens
    compression_metrics["overall"] = {"original_tokens": overall_orig, "after_anchor_tokens": overall_anchor, "after_inverted_index_tokens": overall_index, "after_format_tokens": overall_format, "final_tokens": overall_final, "overall_ratio": (overall_orig / overall_final) if overall_final else 0}
    full_encoding = {"file_name": os.path.basename(excel_path), "sheets": sheets_encoding, "compression_metrics": compression_metrics}
    if output_path:
        with open(output_path, "w", encoding="utf-8") as f:
            json.dump(full_encoding, f, ensure_ascii=False, indent=2)
        logger.info(f"Saved encoding to {output_path}")
    return full_encoding

def extract_cells_near_anchors(sheet, row_anchors, col_anchors, k):
    rows_to_keep = set()
    cols_to_keep = set()
    for r in row_anchors:
        for i in range(max(1, r - k), min(sheet.max_row + 1, r + k + 1)):
            rows_to_keep.add(i)
    for c in col_anchors:
        for i in range(max(1, c - k), min(sheet.max_column + 1, c + k + 1)):
            cols_to_keep.add(i)
    return sorted(list(rows_to_keep)), sorted(list(cols_to_keep))

def cluster_numeric_ranges(sheet, format_map):
    numeric_map = {fmt: cells for fmt, cells in format_map.items() if json.loads(fmt).get("inferred_data_type") == "numeric"}
    if not numeric_map:
        for r in range(1, sheet.max_row + 1):
            for c in range(1, sheet.max_column + 1):
                cell = sheet.cell(row=r, column=c)
                if infer_cell_data_type(cell) == "numeric":
                    fmt_key = json.dumps({"type": detect_semantic_type(cell), "nfs": get_number_format_string(cell)}, sort_keys=True)
                    numeric_map.setdefault(fmt_key, []).append(f"{get_column_letter(c)}{r}")
    return aggregate_formats(sheet, numeric_map)

# -----------------------
# CSV -> XLSX helper
# -----------------------
def convert_csv_to_xlsx(csv_path, xlsx_path):
    if pd is None:
        raise RuntimeError("pandas is required to convert CSV to XLSX (install with `pip install pandas openpyxl`).")
    # try common encodings/separators more simply
    tried = []
    for enc in ['utf-16', 'utf-8', 'latin1']:
        for sep in ['\t', ',', ';', '|']:
            try:
                df = pd.read_csv(csv_path, encoding=enc, sep=sep)
                if df.shape[1] > 1:
                    df.to_excel(xlsx_path, index=False, engine='openpyxl')
                    logger.info(f"Converted CSV -> XLSX (encoding={enc} sep={sep})")
                    return True
            except Exception as e:
                tried.append((enc, sep, str(e)))
                continue
    # fallback
    try:
        df = pd.read_csv(csv_path, encoding='utf-8', sep=None, engine='python')
        df.to_excel(xlsx_path, index=False, engine='openpyxl')
        logger.info("Converted CSV -> XLSX using python engine")
        return True
    except Exception as e:
        logger.error("CSV->XLSX conversion failed. Tried: %s", tried[:5])
        raise

# -----------------------
# CLI entry point
# -----------------------
def main():
    parser = argparse.ArgumentParser(description="SpreadsheetLLM runner (single-file)")
    parser.add_argument("input_file", help="Path to input XLSX or CSV")
    parser.add_argument("-o", "--output", help="Output JSON file path", default=None)
    parser.add_argument("--k", type=int, default=2, help="Neighborhood parameter")
    args = parser.parse_args()

    input_path = Path(args.input_file)
    if not input_path.exists():
        raise SystemExit(f"Input not found: {input_path}")

    # if CSV -> convert
    work_xlsx = str(input_path)
    if input_path.suffix.lower() == ".csv":
        work_xlsx = str(input_path.with_suffix(".xlsx"))
        convert_csv_to_xlsx(str(input_path), work_xlsx)

    if args.output:
        outp = args.output
    else:
        outp = str(Path(work_xlsx).with_suffix("_spreadsheetllm.json"))

    result = spreadsheet_llm_encode_with_helpers(work_xlsx, outp, k=args.k)
    logger.info("Done. Output saved to %s", outp)
    print(json.dumps(result["compression_metrics"]["overall"], indent=2, ensure_ascii=False))

if __name__ == "__main__":
    main()
