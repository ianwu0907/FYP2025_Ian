"""
从 FeTaQA jsonl 文件中随机抽取 200 张表，
同时生成对应的 xlsx 文件和 QA pair，合并进你的自定义 dataset。

用法：
    python fetaqa_extract.py \
        --fetaqa    ./fetaQA-v1_dev.jsonl \
        --custom    ./merged_output_filtered.json \
        --output    ./merged_output_fetaqa.json \
        --xlsx_dir  ./fetaqa_tables \
        --count     200

依赖：
    pip install openpyxl
"""

import json
import os
import re
import random
import argparse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def parse_args():
    parser = argparse.ArgumentParser()
    parser.add_argument("--fetaqa",   required=True, help="FeTaQA jsonl 文件路径")
    parser.add_argument("--custom",   required=True, help="你的自定义 dataset JSON 文件")
    parser.add_argument("--output",   required=True, help="合并后输出的 JSON 文件路径")
    parser.add_argument("--xlsx_dir", required=True, help="xlsx 表格输出目录")
    parser.add_argument("--count",    type=int, default=200, help="抽取表格数量（默认 200）")
    parser.add_argument("--seed",     type=int, default=42,  help="随机种子（默认 42）")
    return parser.parse_args()


def save_xlsx(table_array, filepath):
    """将二维数组保存为 xlsx，第一行作为表头"""
    wb = Workbook()
    ws = wb.active

    header_fill = PatternFill("solid", fgColor="BDD7EE")
    header_font = Font(name="Arial", bold=True, size=10)
    body_font   = Font(name="Arial", size=10)
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    thin  = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for r_idx, row in enumerate(table_array):
        for c_idx, val in enumerate(row):
            cell = ws.cell(row=r_idx + 1, column=c_idx + 1, value=val if val != "" else None)
            cell.border = border
            if r_idx == 0:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center_align
            else:
                cell.font = body_font
                cell.alignment = left_align

    # 自动列宽
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 4, 40)

    wb.save(filepath)


def get_current_max_id(qa_pairs):
    max_num = 0
    for pair in qa_pairs:
        m = re.search(r"(\d+)$", str(pair.get("id", "")))
        if m:
            max_num = max(max_num, int(m.group(1)))
    return max_num


def main():
    args = parse_args()
    os.makedirs(args.xlsx_dir, exist_ok=True)

    # ── 1. 读取 FeTaQA jsonl ──────────────────────────────────────
    print(f"\n📋 读取 FeTaQA 文件: {args.fetaqa}")
    samples = []
    with open(args.fetaqa, encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if line:
                samples.append(json.loads(line))
    print(f"   共 {len(samples)} 条样本")

    # ── 2. 随机抽取 ──────────────────────────────────────────────
    random.seed(args.seed)
    count = min(args.count, len(samples))
    selected = random.sample(samples, count)
    print(f"   随机抽取: {count} 条 (seed={args.seed})\n")

    # ── 3. 加载自定义 dataset ────────────────────────────────────
    with open(args.custom, encoding="utf-8") as f:
        custom = json.load(f)
    existing_pairs = custom["qa_pairs"]
    max_id = get_current_max_id(existing_pairs)
    print(f"📦 现有 QA pairs: {len(existing_pairs)} 条")

    # ── 4. 生成 xlsx + QA pair ───────────────────────────────────
    new_pairs = []
    saved_xlsx = 0
    failed = []

    for sample in selected:
        feta_id     = sample.get("feta_id", "")
        table_array = sample.get("table_array", [])
        question    = sample.get("question", "")
        answer      = sample.get("answer", "")
        page_title  = sample.get("table_page_title", "")
        sec_title   = sample.get("table_section_title", "")

        if not table_array or not question or not answer:
            failed.append(feta_id)
            continue

        # 文件名用 feta_id
        xlsx_name = f"fetaqa_{feta_id}.xlsx"
        xlsx_path = os.path.join(args.xlsx_dir, xlsx_name)

        try:
            save_xlsx(table_array, xlsx_path)
            saved_xlsx += 1
        except Exception as e:
            print(f"   ⚠️  生成 xlsx 失败 feta_id={feta_id}: {e}")
            failed.append(feta_id)
            continue

        max_id += 1
        new_pairs.append({
            "id": f"qa-{max_id:03d}",
            "table_file": xlsx_name,
            "question": question,
            "answers": [answer],
            "answer_type": "text",
            "source": "fetaqa",
            "fetaqa_id": str(feta_id),
            "table_page_title": page_title,
            "table_section_title": sec_title
        })

    print(f"\n✅ 结果:")
    print(f"   成功生成 xlsx: {saved_xlsx} 张 → {args.xlsx_dir}")
    if failed:
        print(f"   跳过（数据缺失）: {len(failed)} 条")

    # ── 5. 合并并保存 ────────────────────────────────────────────
    merged = existing_pairs + new_pairs
    custom["qa_pairs"] = merged
    custom["metadata"]["total_qa_pairs"] = len(merged)

    with open(args.output, "w", encoding="utf-8") as f:
        json.dump(custom, f, ensure_ascii=False, indent=2)

    print(f"\n🎉 合并完成:")
    print(f"   新增 {len(new_pairs)} 条 FeTaQA QA pairs")
    print(f"   总计 {len(merged)} 条 → {args.output}")


if __name__ == "__main__":
    main()
