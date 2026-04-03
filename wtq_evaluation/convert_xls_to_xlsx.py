#!/usr/bin/env python3
"""将 tables_for_dataset/ 里的 .xls 文件原地转换为 .xlsx，原 .xls 保留备份。"""

from pathlib import Path
import xlrd
import openpyxl

INPUT_DIR = Path(__file__).parent / "tables_for_dataset"

def convert(xls_path: Path):
    xlsx_path = xls_path.with_suffix(".xlsx")
    if xlsx_path.exists():
        print(f"  SKIP (already exists): {xlsx_path.name}")
        return

    wb_new = openpyxl.Workbook()
    wb_new.remove(wb_new.active)  # remove default sheet

    wb_old = xlrd.open_workbook(str(xls_path), formatting_info=False)
    for sheet_name in wb_old.sheet_names():
        ws_old = wb_old.sheet_by_name(sheet_name)
        ws_new = wb_new.create_sheet(title=sheet_name)
        for row_idx in range(ws_old.nrows):
            for col_idx in range(ws_old.ncols):
                cell = ws_old.cell(row_idx, col_idx)
                ws_new.cell(row=row_idx + 1, column=col_idx + 1, value=cell.value)

    wb_new.save(str(xlsx_path))
    print(f"  OK: {xls_path.name} -> {xlsx_path.name}")

def main():
    xls_files = sorted(INPUT_DIR.glob("*.xls"))
    print(f"Found {len(xls_files)} .xls file(s) in {INPUT_DIR}\n")
    ok, fail = 0, 0
    for f in xls_files:
        try:
            convert(f)
            ok += 1
        except Exception as e:
            print(f"  FAIL: {f.name} -> {e}")
            fail += 1
    print(f"\nDone: {ok} converted, {fail} failed")

if __name__ == "__main__":
    main()
